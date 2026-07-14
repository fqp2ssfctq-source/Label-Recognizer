"""
Optical Transceiver Label Inspector  v2.0
==========================================
각 라벨 구역을 개별 crop 후 OCR / 바코드 인식.
템플릿에서 구역(rect)을 드래그로 정의 → 검사 시 해당 구역만 처리 → 빠름(2초 이내).

의존 패키지:
  pip install opencv-python pillow pytesseract pyzbar numpy openpyxl

tesseract-ocr 설치 필요:
  Windows : https://github.com/UB-Mannheim/tesseract/wiki
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import datetime
import json
import os
import csv

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

import cv2
import numpy as np
from PIL import Image, ImageTk
import pytesseract

_TESS_DEFAULT = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
if os.path.exists(_TESS_DEFAULT):
    pytesseract.pytesseract.tesseract_cmd = _TESS_DEFAULT

try:
    from pyzbar import pyzbar
    PYZBAR_OK = True
except ImportError:
    PYZBAR_OK = False

try:
    import tesserocr
    from tesserocr import PSM
    TESSEROCR_OK = True
except ImportError:
    TESSEROCR_OK = False

# ──────────────────────────────────────────────
CLR = {
    "bg":        "#1a1d23",
    "panel":     "#22262f",
    "border":    "#2e3340",
    "accent":    "#00c2ff",
    "ok":        "#00e676",
    "ng":        "#ff3d3d",
    "warn":      "#ffaa00",
    "text":      "#e8ecf0",
    "subtext":   "#8899aa",
    "btn":       "#2c3244",
    "btn_hover": "#3a4257",
    "btn_on":    "#1a5f7a",
}

# 구역 유형별 색상 (hex) — 마법사와 캔버스 오버레이에서 공유
REGION_COL = {
    "desc":    "#3399ff",
    "pn":      "#22bb55",
    "sn":      "#ff9900",
    "ocr":     "#cc44ff",
    "barcode": "#ff4444",
    "ci":      "#00cccc",
    "class1":  "#ff66cc",
    "certi":   "#99cc00",
}
REGION_COL_EXTRA = ["#ff6633", "#6699ff", "#cc9900", "#33cc99",
                    "#cc6633", "#9966ff", "#33aacc", "#ff9966"]

# 그림(로고/마크) 구역 — OCR 대신 기준 이미지와 이미지 유사도로 비교
IMAGE_REGION_TYPES = {"ci", "certi"}

def _hex_to_rgb(h: str):
    """#RRGGBB → (R,G,B) for cv2 (RGB 이미지 기준)."""
    h = h.lstrip("#")
    return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))

FONT_TITLE  = ("Consolas", 13, "bold")
FONT_BODY   = ("Consolas", 10)
FONT_SMALL  = ("Consolas",  9)
FONT_RESULT = ("Consolas", 11, "bold")


# ──────────────────────────────────────────────
class LabelAnalyzer:
    """라벨 이미지 처리 유틸리티. 모두 정적 메서드."""

    @staticmethod
    def preprocess_crop(bgr):
        """OCR 전처리: 높이 최소 80px 보장 → OTSU 이진화 → 여백 추가."""
        h, w = bgr.shape[:2]
        # 높이 기준 업스케일 (문자가 너무 작으면 OCR 실패)
        if h < 80:
            scale = 80 / h
            bgr = cv2.resize(bgr, (int(w * scale), int(h * scale)),
                             interpolation=cv2.INTER_CUBIC)
            h, w = bgr.shape[:2]
        # 너비가 작으면 추가 확대
        if w < 300:
            scale = 300 / w
            bgr = cv2.resize(bgr, (int(w * scale), int(h * scale)),
                             interpolation=cv2.INTER_CUBIC)
            h, w = bgr.shape[:2]
        # 너무 크면 축소
        if w > 1400:
            bgr = cv2.resize(bgr, (1400, int(h * 1400 / w)),
                             interpolation=cv2.INTER_AREA)
        gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
        gray = cv2.GaussianBlur(gray, (3, 3), 0)
        # OTSU: 인쇄된 밝은 배경 + 어두운 문자에 최적
        _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        # 배경이 어두우면 반전 (흰 바탕 보장)
        if np.mean(binary) < 128:
            binary = cv2.bitwise_not(binary)
        # Tesseract가 가장자리 문자를 잘 인식하도록 흰 여백 추가
        return cv2.copyMakeBorder(binary, 12, 12, 12, 12,
                                  cv2.BORDER_CONSTANT, value=255)

    @staticmethod
    def _postprocess_ocr(text, typ):
        """OCR 결과 후처리: 불필요한 접두사 제거 및 정리."""
        import re
        text = re.sub(r"\s+", " ", text).strip()
        if typ in ("sn", "pn"):
            # "S/N :", "P/N:", "SN:", "PN:", "S/N -" 등 접두사 제거
            text = re.sub(r"^[SP][/\.]?[NP]\s*[:\-\s]+", "", text, flags=re.IGNORECASE)
            text = text.strip("'\"` ").strip()
        return text

    @staticmethod
    def ocr_crop(bgr, typ="ocr"):
        """crop 이미지에서 OCR 텍스트 추출. typ에 따라 PSM·문자셋 최적화."""
        proc = LabelAnalyzer.preprocess_crop(bgr)

        # 유형별 PSM: 단일 코드(P/N, S/N)는 PSM 7(한 줄), 나머지는 PSM 6(블록)
        _PSM_STR = {
            "pn": "7", "sn": "7", "desc": "7",
            "class1": "7", "ci": "8", "certi": "6",
        }
        psm_str = _PSM_STR.get(typ, "6")

        # 코드 유형은 문자 제한으로 오인식 감소
        _WHITELIST = {
            "pn": "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-/. ",
            "sn": "ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-/: ",
        }
        wl = _WHITELIST.get(typ, "")

        if TESSEROCR_OK:
            try:
                _psm_map = {
                    "6": PSM.SINGLE_BLOCK,
                    "7": PSM.SINGLE_LINE,
                    "8": PSM.SINGLE_WORD,
                }
                pil  = Image.fromarray(proc)
                _psm = _psm_map.get(psm_str, PSM.SINGLE_BLOCK)
                if wl:
                    # 문자 화이트리스트 적용 (P/N·S/N 숫자/문자 오인식 감소)
                    with tesserocr.PyTessBaseAPI(lang="eng", psm=_psm) as api:
                        api.SetVariable("tessedit_char_whitelist", wl)
                        api.SetImage(pil)
                        text = api.GetUTF8Text().strip()
                else:
                    text = tesserocr.image_to_text(pil, lang="eng", psm=_psm).strip()
                return LabelAnalyzer._postprocess_ocr(text, typ)
            except Exception:
                pass

        cfg = f"--oem 3 --psm {psm_str}"
        if wl:
            cfg += f" -c tessedit_char_whitelist={wl}"
        try:
            text = pytesseract.image_to_string(proc, config=cfg, timeout=8).strip()
            return LabelAnalyzer._postprocess_ocr(text, typ)
        except Exception:
            return ""

    # OpenCV 검출기 지연 초기화 캐시 (생성 비용 1회)
    _bardet      = None
    _bardet_init = False
    _qrdet       = None
    _qrdet_init  = False

    @staticmethod
    def _get_bardet():
        if not LabelAnalyzer._bardet_init:
            LabelAnalyzer._bardet_init = True
            try:
                LabelAnalyzer._bardet = cv2.barcode.BarcodeDetector()
            except Exception:
                LabelAnalyzer._bardet = None
        return LabelAnalyzer._bardet

    @staticmethod
    def _get_qrdet():
        if not LabelAnalyzer._qrdet_init:
            LabelAnalyzer._qrdet_init = True
            try:
                LabelAnalyzer._qrdet = cv2.QRCodeDetector()
            except Exception:
                LabelAnalyzer._qrdet = None
        return LabelAnalyzer._qrdet

    @staticmethod
    def decode_barcodes(bgr):
        """crop에서 바코드/QR 디코딩.
        다중 엔진(pyzbar + cv2.barcode + cv2.QR) × 다중 전처리 × 업스케일/세로확장
        × 회전 재시도. 얇은 1D 바코드(낮은 높이)도 세로로 늘려 스캔 라인 확보."""
        if bgr is None or getattr(bgr, "size", 0) == 0:
            return []

        # ── 엔진 ─────────────────────────────────
        def _eng_pyzbar(gray):
            if not PYZBAR_OK:
                return []
            try:
                codes = pyzbar.decode(gray)
            except Exception:
                return []
            return [{"type": o.type,
                     "data": o.data.decode("utf-8", errors="replace")}
                    for o in codes]

        def _eng_cv2bar(gray):
            det = LabelAnalyzer._get_bardet()
            if det is None:
                return []
            try:
                ok, info, types, _ = det.detectAndDecodeWithType(gray)
            except Exception:
                return []
            if not ok:
                return []
            return [{"type": t or "BARCODE", "data": s}
                    for s, t in zip(info, types) if s]

        def _eng_cv2qr(gray):
            det = LabelAnalyzer._get_qrdet()
            if det is None:
                return []
            try:
                s, _pts, _ = det.detectAndDecode(gray)
            except Exception:
                return []
            return [{"type": "QRCODE", "data": s}] if s else []

        # ── 전처리 후보 (pyzbar용) ───────────────
        def _candidates(gray):
            imgs = [gray]
            _, th = cv2.threshold(gray, 0, 255,
                                  cv2.THRESH_BINARY | cv2.THRESH_OTSU)
            imgs.append(th)
            imgs.append(cv2.bitwise_not(th))               # 배경 어두울 때
            imgs.append(cv2.adaptiveThreshold(
                gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                cv2.THRESH_BINARY, 15, 4))
            imgs.append(cv2.createCLAHE(2.0, (4, 4)).apply(gray))
            # 언샤프 마스킹 (블러 보정)
            blur = cv2.GaussianBlur(gray, (0, 0), 1.5)
            imgs.append(cv2.addWeighted(gray, 1.6, blur, -0.6, 0))
            return imgs

        # ── pyzbar용 스케일 (가로 업스케일 + 얇은 바코드 세로 확장) ──
        def _scaled_grays(gray):
            h, w = gray.shape[:2]
            outs = []
            if w < 800:
                f = 3 if w < 300 else 2
                outs.append(cv2.resize(gray, (w * f, h * f),
                                       interpolation=cv2.INTER_CUBIC))
            outs.append(gray)
            if h < 60:                                     # 얇은 1D 바코드
                vf = max(2, 60 // max(1, h))
                outs.append(cv2.resize(gray, (w, h * vf),
                                       interpolation=cv2.INTER_NEAREST))
                if w < 800:
                    outs.append(cv2.resize(gray, (w * 2, h * vf),
                                           interpolation=cv2.INTER_CUBIC))
            return outs

        # ── cv2.barcode용 스케일 (입력 크기와 무관한 목표 너비) ──
        #    바가 얇은(저해상도) 바코드는 특정 배율에서만 검출되므로 넓게 스윕.
        def _bar_scales(gray):
            h, w = gray.shape[:2]
            outs = []
            for tw in (600, 900, 1200, 1600, 2000):
                if tw <= w:
                    continue
                f = tw / w
                outs.append(cv2.resize(gray, (int(w * f), int(h * f)),
                                       interpolation=cv2.INTER_CUBIC))
            outs.append(gray)
            return outs

        # ── 회전 (정방향 → 180 → 90 → 270) ───────
        def _rotations(img):
            yield img
            yield cv2.rotate(img, cv2.ROTATE_180)
            yield cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
            yield cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)

        PAD = 20  # quiet zone 여백
        # cv2.barcode(1D)는 저품질 영상에서 배율마다 다른 값을 내는
        # 오인식 경향이 있어 곧바로 신뢰하지 않고 득표 후 교차검증한다.
        bar_votes = {}
        for rot in _rotations(bgr):
            padded = cv2.copyMakeBorder(rot, PAD, PAD, PAD, PAD,
                                        cv2.BORDER_CONSTANT, value=(255, 255, 255))
            gray0 = (cv2.cvtColor(padded, cv2.COLOR_BGR2GRAY)
                     if padded.ndim == 3 else padded)

            # 1) pyzbar — 체크섬을 엄격히 검증 → 적중 시 즉시 신뢰
            for g in _scaled_grays(gray0):
                for cand in _candidates(g):
                    r = _eng_pyzbar(cand)
                    if r:
                        return r
            # 2) cv2.QR — 오류정정 내장 → 적중 시 즉시 신뢰
            for g in _scaled_grays(gray0):
                r = _eng_cv2qr(g)
                if r:
                    return r
            # 3) cv2.barcode — 여러 배율에서 득표만 수집 (자체 이진화 → 원본 그레이)
            for g in _bar_scales(gray0):
                for item in _eng_cv2bar(g):
                    key = (item["type"], item["data"])
                    bar_votes[key] = bar_votes.get(key, 0) + 1

        # cv2.barcode 결과는 서로 다른 배율에서 2회 이상 동일하게 나올 때만 채택
        for (typ, data), n in sorted(bar_votes.items(), key=lambda kv: -kv[1]):
            if n >= 2:
                return [{"type": typ, "data": data}]
        return []


# ──────────────────────────────────────────────
class TemplateManager:
    """
    템플릿 JSON 포맷:
    {
      "template_name": "...",
      "img_size": [w, h],          // 템플릿 생성 시 이미지 크기
      "ref_image": "name_ref.jpg", // 라벨 추적용 참조 이미지 (선택)
      "regions": [
        {"label": "P/N", "type": "pn",      "rect": [x,y,w,h], "text": "ZB7784099-DIL"},
        {"label": "S/N", "type": "sn",      "rect": [x,y,w,h], "text": ""},
        {"label": "바코드", "type": "barcode","rect": [x,y,w,h], "text": ""},
        ...
      ]
    }
    type 종류: barcode | pn | sn | desc | ocr | cust_*
    text: 기대값(비어있으면 존재 여부만 확인)
    """

    def __init__(self):
        self.data     = None
        self.path     = ""
        self.ref_bgr  = None       # 참조 이미지 (ORB 추적용)
        self._orb_kp  = None
        self._orb_des = None
        self._orb     = cv2.ORB_create(1000)
        self._matcher = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)

    @property
    def loaded(self):
        return self.data is not None

    @property
    def name(self):
        return self.data["template_name"] if self.loaded else "템플릿 없음"

    @property
    def has_tracking(self):
        """참조 이미지가 있고 ORB 특징점이 충분하면 True."""
        return self._orb_des is not None and len(self._orb_des) >= 20

    def load(self, path):
        with open(path, encoding="utf-8") as f:
            self.data = json.load(f)
        self.path = path
        self._load_ref_image()

    def _load_ref_image(self):
        """ref_image 경로의 참조 이미지 로드 및 ORB 특징점 계산."""
        self.ref_bgr = None
        self._orb_kp = None
        self._orb_des = None
        ref = self.data.get("ref_image", "") if self.data else ""
        if not ref:
            return
        if not os.path.isabs(ref):
            ref = os.path.join(os.path.dirname(self.path), ref)
        if not os.path.exists(ref):
            return
        bgr = cv2.imread(ref)
        if bgr is None:
            return
        self.ref_bgr = bgr
        # 절반 해상도에서 특징점 계산 (속도)
        small = cv2.resize(bgr, (bgr.shape[1] // 2, bgr.shape[0] // 2))
        gray  = cv2.cvtColor(small, cv2.COLOR_BGR2GRAY)
        self._orb_kp, self._orb_des = self._orb.detectAndCompute(gray, None)

    def match_frame(self, bgr):
        """현재 프레임에서 참조 이미지 위치를 찾아 호모그래피 반환. 실패 시 None."""
        if not self.has_tracking:
            return None
        h, w = bgr.shape[:2]
        cur_s = cv2.resize(bgr, (w // 2, h // 2))
        cur_g = cv2.cvtColor(cur_s, cv2.COLOR_BGR2GRAY)
        kp, des = self._orb.detectAndCompute(cur_g, None)
        if des is None or len(des) < 10:
            return None
        matches = self._matcher.match(self._orb_des, des)
        matches = sorted(matches, key=lambda m: m.distance)
        good    = [m for m in matches if m.distance < 60][:60]
        if len(good) < 12:
            return None
        src_pts = np.float32([self._orb_kp[m.queryIdx].pt for m in good]).reshape(-1, 1, 2)
        dst_pts = np.float32([kp[m.trainIdx].pt           for m in good]).reshape(-1, 1, 2)
        H, mask = cv2.findHomography(src_pts, dst_pts, cv2.RANSAC, 5.0)
        if H is None or int(mask.sum()) < 10:
            return None
        # H는 절반 해상도 공간에서 ref→cur 변환
        # H_full = S_inv @ H @ S  (S = diag(0.5, 0.5, 1) 로 full→half 스케일)
        S     = np.diag([0.5, 0.5, 1.0])
        S_inv = np.diag([2.0, 2.0, 1.0])
        return S_inv @ H @ S

    @staticmethod
    def transform_corners(H, rect):
        """호모그래피 H로 [x,y,w,h] → 변환된 4개 꼭짓점 [[x,y], ...] (기울기 보존)."""
        rx, ry, rw, rh = rect
        corners = np.float32([
            [rx,      ry     ],
            [rx + rw, ry     ],
            [rx + rw, ry + rh],
            [rx,      ry + rh],
        ]).reshape(-1, 1, 2)
        return cv2.perspectiveTransform(corners, H).reshape(-1, 2)

    @staticmethod
    def transform_rect(H, rect):
        """호모그래피 H로 [x,y,w,h] → 축 정렬 바운딩박스 [x,y,w,h] (폴백용)."""
        t = TemplateManager.transform_corners(H, rect)
        nx = int(t[:, 0].min());  ny = int(t[:, 1].min())
        nw = max(1, int(t[:, 0].max()) - nx)
        nh = max(1, int(t[:, 1].max()) - ny)
        return [nx, ny, nw, nh]

    def save(self, path):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)
        self.path = path

    def inspect(self, bgr, H=None):
        """
        구역별 검사. H(호모그래피)가 있으면 라벨 위치 추적 적용.
        반환: [{"label", "type", "ok", "value", "reason"}, ...]
        """
        if not self.loaded:
            return []
        ih, iw = bgr.shape[:2]
        if H is None:
            tw, th = self.data.get("img_size", [iw, ih])
            sx = iw / tw if tw else 1.0
            sy = ih / th if th else 1.0

        import threading as _th
        regions = self.data.get("regions", [])
        results = [None] * len(regions)

        def _run(i, reg):
            if H is not None:
                # 기울어진 4점 → perspective warp로 정렬된 crop 추출
                pts = self.transform_corners(H, reg["rect"]).astype(np.float32)
                w_out = int(max(np.linalg.norm(pts[1] - pts[0]),
                                np.linalg.norm(pts[2] - pts[3])))
                h_out = int(max(np.linalg.norm(pts[3] - pts[0]),
                                np.linalg.norm(pts[2] - pts[1])))
                w_out = max(1, w_out);  h_out = max(1, h_out)
                dst = np.float32([[0, 0], [w_out, 0], [w_out, h_out], [0, h_out]])
                M   = cv2.getPerspectiveTransform(pts, dst)
                crop = cv2.warpPerspective(bgr, M, (w_out, h_out))
            else:
                r = reg["rect"]
                rx, ry, rw, rh = int(r[0]*sx), int(r[1]*sy), int(r[2]*sx), int(r[3]*sy)
                x0 = max(0, rx);         y0 = max(0, ry)
                x1 = min(iw, rx + rw);   y1 = min(ih, ry + rh)
                if x1 <= x0 or y1 <= y0:
                    results[i] = {"label": reg.get("label", reg["type"]),
                                   "type": reg["type"], "ok": False,
                                   "value": "", "reason": "구역 범위 오류"}
                    return
                crop = bgr[y0:y1, x0:x1]
            results[i] = self._check(crop, reg)

        threads = [_th.Thread(target=_run, args=(i, reg), daemon=True)
                   for i, reg in enumerate(regions)]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        self._cross_check_barcodes(regions, results)
        return results

    @staticmethod
    def _barcode_match_target(reg_label):
        """'S/N bar code' → 'S/N' 처럼 짝이 되는 텍스트 구역 라벨을 추출. 없으면 None."""
        import re
        base = re.sub(r"\s*(bar\s*code|barcode|바코드)\s*$", "",
                      reg_label or "", flags=re.IGNORECASE).strip()
        if base and base.lower() != (reg_label or "").strip().lower():
            return base
        return None

    def _cross_check_barcodes(self, regions, results):
        """바코드-텍스트 교차검증: 'X bar code' 바코드는 'X' 텍스트 구역 값과
        동일하면 OK, 미인식·불일치면 NG. (짝 텍스트 구역이 있을 때만 적용)"""
        text_by_label = {}
        for r in results:
            if r and r.get("type") != "barcode":
                text_by_label[(r.get("label") or "").strip().lower()] = r
        for i, reg in enumerate(regions):
            if reg.get("type") != "barcode":
                continue
            res = results[i]
            if not res:
                continue
            base = self._barcode_match_target(reg.get("label", ""))
            if not base:
                continue  # 짝 없음 → 기존 고정 기대텍스트 방식 유지
            sib = text_by_label.get(base.lower())
            if sib is None:
                continue
            decoded = (res.get("value") or "").strip()
            target  = (sib.get("value") or "").strip()
            norm = lambda s: s.replace(" ", "").replace("-", "").upper()
            if not decoded:
                res["ok"] = False
                res["reason"] = "바코드 미인식"
            elif target and norm(decoded) == norm(target):
                res["ok"] = True
                res["reason"] = ""
                res["value"] = f"{decoded} (= {sib.get('label')})"
            elif target:
                res["ok"] = False
                res["reason"] = f"{sib.get('label')}({target})와 불일치"
            else:
                # 짝 텍스트가 비어 있으면 비교 불가 → 바코드 존재만으로 판단
                res["ok"] = bool(decoded)
                res["reason"] = "" if decoded else "바코드 미인식"

    def _ref_region_crop(self, reg):
        """기준 이미지(ref_bgr)에서 해당 구역을 잘라 반환. 없으면 None."""
        ref = getattr(self, "ref_bgr", None)
        if ref is None:
            return None
        rh_, rw_ = ref.shape[:2]
        img_size = self.data.get("img_size") if self.data else None
        if img_size and img_size[0] and img_size[1]:
            sx = rw_ / img_size[0];  sy = rh_ / img_size[1]
        else:
            sx = sy = 1.0
        rx, ry, rw, rh = reg["rect"]
        x0 = max(0, int(rx * sx));        y0 = max(0, int(ry * sy))
        x1 = min(rw_, int((rx + rw) * sx)); y1 = min(rh_, int((ry + rh) * sy))
        if x1 <= x0 or y1 <= y0:
            return None
        return ref[y0:y1, x0:x1]

    @staticmethod
    def _image_similarity(a, b):
        """두 이미지의 정규화 상관계수 유사도 (-1~1, 클수록 유사). 조명 보정 포함."""
        if a is None or b is None or a.size == 0 or b.size == 0:
            return -1.0
        ga = cv2.cvtColor(a, cv2.COLOR_BGR2GRAY) if a.ndim == 3 else a
        gb = cv2.cvtColor(b, cv2.COLOR_BGR2GRAY) if b.ndim == 3 else b
        W, H = 192, 96
        ga = cv2.equalizeHist(cv2.resize(ga, (W, H)))
        gb = cv2.equalizeHist(cv2.resize(gb, (W, H)))
        res = cv2.matchTemplate(ga, gb, cv2.TM_CCOEFF_NORMED)
        return float(res[0, 0])

    def _check_image(self, crop, reg, label):
        """그림 구역 검사: 기준 이미지와 유사도 비교."""
        ref_crop = self._ref_region_crop(reg)
        if ref_crop is None:
            return {"label": label, "type": reg["type"], "ok": False,
                    "value": "", "reason": "기준 이미지 없음 (템플릿 재저장 필요)"}
        sim = self._image_similarity(crop, ref_crop)
        thr = reg.get("img_thr", 0.40)   # 유사도 임계값 (구역별 조정 가능)
        ok  = sim >= thr
        return {"label": label, "type": reg["type"], "ok": ok,
                "value": f"유사도 {sim:.2f}",
                "reason": "" if ok else f"기준 이미지와 불일치 ({sim:.2f} < {thr:.2f})"}

    def _check(self, crop, reg):
        typ      = reg["type"]
        label    = reg.get("label") or typ
        expected = reg.get("text", "").strip()

        # 그림(로고/마크) 구역 — 이미지 유사도 비교
        if typ in IMAGE_REGION_TYPES or reg.get("compare") == "image":
            return self._check_image(crop, reg, label)

        if typ == "barcode":
            codes = LabelAnalyzer.decode_barcodes(crop)
            ok    = bool(codes)
            value = codes[0]["data"] if codes else ""
            reason = "" if ok else "바코드 미인식"
            if ok and expected and value != expected:
                ok = False
                reason = f"불일치 (기대: {expected})"
            return {"label": label, "type": typ, "ok": ok,
                    "value": value, "reason": reason}

        # OCR 기반 구역 (pn / sn / desc / ocr / cust_*)
        text = LabelAnalyzer.ocr_crop(crop, typ=typ)
        if expected:
            ok     = expected.lower() in text.lower()
            reason = "" if ok else f'"{expected}" 미발견'
        else:
            ok     = bool(text.strip())
            reason = "" if ok else "텍스트 없음"
        return {"label": label, "type": typ, "ok": ok,
                "value": text.strip(), "reason": reason}


# ──────────────────────────────────────────────
class App(tk.Tk):
    PREVIEW_W  = 560
    PREVIEW_H  = 360
    _ZOOM_MIN  = 1.0
    _ZOOM_MAX  = 4.0
    _ZOOM_STEP = 0.5
    _CSV_COLS  = ["Timestamp", "Verdict", "Template",
                  "Label", "Type", "OK", "Value", "Reason"]

    def __init__(self):
        super().__init__()
        self.title("Optical Transceiver Label Inspector  v2.0")
        self.configure(bg=CLR["bg"])
        self.resizable(True, True)

        self.tmpl            = TemplateManager()
        self.cap             = None
        self.cam_running     = False
        self._latest_frame   = None   # 카메라 리더 → 디스플레이 공유 프레임
        self._frame_count    = 0
        self.current_bgr     = None
        self._static_bgr     = None
        self.history         = []
        self._after_id       = None
        self._rotate         = 0
        self._zoom           = 1.0
        self._pan_x          = 0
        self._pan_y          = 0
        self._drag_start     = None
        self._cam_fmt        = "YUY2"
        self._cam_res        = (640, 480)
        self._current_cam_idx = 0
        self._auto_inspect   = False
        self._auto_interval  = 5
        self._auto_after_id  = None
        self._track_H        = None   # 현재 호모그래피 (None=추적 없음)
        self._track_ok       = False  # 추적 성공 여부
        self._tracking       = False  # 추적 스레드 실행 플래그

        _dir = os.path.dirname(os.path.abspath(__file__))
        self._app_dir     = _dir   # GUI 위치 폴더 (템플릿 기본 경로)
        self._log_path    = os.path.join(
            _dir, f"label_log_{datetime.datetime.now().strftime('%Y%m%d')}.csv"
        )
        self._config_path = os.path.join(_dir, ".label_config.json")
        self._ensure_csv_header()
        self._build_ui()
        self._load_config()
        self.bind("<space>", lambda e: self._inspect()
                  if not isinstance(self.focus_get(), (tk.Entry, tk.Spinbox)) else None)
        self.bind("<F5>", lambda e: self._inspect())
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    # ── UI ──────────────────────────────────────
    def _build_ui(self):
        self.columnconfigure(0, weight=3)
        self.columnconfigure(1, weight=2)
        self.rowconfigure(1, weight=1)

        hdr = tk.Frame(self, bg=CLR["panel"], height=48)
        hdr.grid(row=0, column=0, columnspan=2, sticky="ew")
        tk.Label(hdr, text="▣  LABEL INSPECTOR", font=("Consolas", 14, "bold"),
                 fg=CLR["accent"], bg=CLR["panel"]).pack(side="left", padx=16, pady=10)
        self._lbl_clock = tk.Label(hdr, text="", font=FONT_SMALL,
                                   fg=CLR["subtext"], bg=CLR["panel"])
        self._lbl_clock.pack(side="right", padx=16)
        self._lbl_tmpl = tk.Label(hdr, text="[ 템플릿 없음 ]", font=FONT_SMALL,
                                  fg=CLR["warn"], bg=CLR["panel"])
        self._lbl_tmpl.pack(side="right", padx=8)
        self._tick_clock()

        # 좌측 패널
        left = tk.Frame(self, bg=CLR["bg"])
        left.grid(row=1, column=0, sticky="nsew", padx=8, pady=8)
        left.rowconfigure(0, weight=1)
        left.columnconfigure(0, weight=1)

        self._canvas = tk.Canvas(left, width=self.PREVIEW_W, height=self.PREVIEW_H,
                                 bg="#0d1117", highlightthickness=1,
                                 highlightbackground=CLR["border"])
        self._canvas.grid(row=0, column=0, sticky="nsew")
        self._canvas.create_text(self.PREVIEW_W // 2, self.PREVIEW_H // 2,
                                 text="[ 카메라 시작 또는 파일 열기 ]",
                                 fill=CLR["subtext"], font=FONT_BODY, tags="hint")
        self._canvas.bind("<MouseWheel>",      self._on_mousewheel)
        self._canvas.bind("<Enter>",           lambda e: self._canvas.focus_set())
        self._canvas.bind("<Button-1>",        self._on_drag_start)
        self._canvas.bind("<B1-Motion>",       self._on_drag_move)
        self._canvas.bind("<ButtonRelease-1>", self._on_drag_end)

        # ctrl1: 카메라 제어
        ctrl1 = tk.Frame(left, bg=CLR["bg"])
        ctrl1.grid(row=1, column=0, sticky="ew", pady=(4, 0))
        self._btn_cam  = self._btn(ctrl1, "📷  카메라 선택", self._show_cam_dialog)
        self._btn_stop = self._btn(ctrl1, "⏹  중지",        self._stop_camera)
        self._btn_file = self._btn(ctrl1, "📂  파일 열기",   self._open_file)
        self._btn_snap = self._btn(ctrl1, "🔍  검사 실행",   self._inspect, accent=True)
        self._auto_spin = tk.Spinbox(
            ctrl1, from_=2, to=60, width=3,
            bg=CLR["btn"], fg=CLR["text"],
            buttonbackground=CLR["panel"],
            insertbackground=CLR["text"],
            relief="flat", font=FONT_BODY, justify="center",
        )
        self._auto_spin.delete(0, "end")
        self._auto_spin.insert(0, "5")
        self._btn_auto = self._btn(ctrl1, "⟳  자동", self._toggle_auto_inspect)
        self._btn_cam.pack(side="left", padx=4, pady=4)
        self._btn_stop.pack(side="left", padx=(0, 8), pady=4)
        self._btn_file.pack(side="left", padx=4, pady=4)
        self._btn_snap.pack(side="left", padx=4, pady=4)
        tk.Frame(ctrl1, bg=CLR["border"], width=1).pack(side="left", fill="y", padx=8, pady=4)
        self._btn_auto.pack(side="left", padx=4, pady=4)
        self._auto_spin.pack(side="left", padx=2, pady=4)
        tk.Label(ctrl1, text="초", font=FONT_SMALL,
                 fg=CLR["subtext"], bg=CLR["bg"]).pack(side="left")
        self._btn_stop.configure(state="disabled")

        # ctrl2: 포맷 / 템플릿 / 회전 / 줌
        ctrl2 = tk.Frame(left, bg=CLR["bg"])
        ctrl2.grid(row=2, column=0, sticky="ew", pady=(0, 4))

        tk.Label(ctrl2, text="포맷:", font=FONT_SMALL,
                 fg=CLR["subtext"], bg=CLR["bg"]).pack(side="left", padx=(4, 2))
        self._fmt_var = tk.StringVar(value="YUY2")
        ttk.Combobox(ctrl2, textvariable=self._fmt_var, values=["MJPG", "YUY2", "AUTO"],
                     width=6, state="readonly", font=FONT_SMALL).pack(side="left", padx=(0, 2))
        tk.Label(ctrl2, text="해상도:", font=FONT_SMALL,
                 fg=CLR["subtext"], bg=CLR["bg"]).pack(side="left", padx=(6, 2))
        self._res_var = tk.StringVar(value="640×480")
        _res_cb = ttk.Combobox(ctrl2, textvariable=self._res_var,
                     values=["640×480", "1280×720", "1920×1080"],
                     width=9, state="readonly", font=FONT_SMALL)
        _res_cb.pack(side="left", padx=(0, 2))
        _res_cb.bind("<<ComboboxSelected>>", lambda e: self._on_res_change())

        tk.Frame(ctrl2, bg=CLR["border"], width=1).pack(side="left", fill="y", padx=6, pady=4)

        self._btn_tmpl     = self._btn(ctrl2, "📋 템플릿",   self._load_template)
        self._btn_new      = self._btn(ctrl2, "✨ 새 템플릿", self._template_wizard)
        self._btn_edit     = self._btn(ctrl2, "✏ 편집",     self._edit_template)
        self._btn_tmpl_del = self._btn(ctrl2, "✕ 템플릿 해제", self._clear_template,
                                       bg=CLR["ng"])
        self._btn_tmpl.pack(side="left",     padx=3, pady=2)
        self._btn_new.pack(side="left",      padx=3, pady=2)
        self._btn_edit.pack(side="left",     padx=3, pady=2)
        self._btn_tmpl_del.pack(side="left", padx=3, pady=2)

        tk.Frame(ctrl2, bg=CLR["border"], width=1).pack(side="left", fill="y", padx=6, pady=4)

        self._btn(ctrl2, "↻ 90°", self._rotate_cw).pack(side="left", padx=3, pady=2)
        self._btn(ctrl2, "↺ 90°", self._rotate_ccw).pack(side="left", padx=3, pady=2)
        self._lbl_rotate = tk.Label(ctrl2, text="0°", font=FONT_SMALL,
                                    fg=CLR["subtext"], bg=CLR["bg"], width=4)
        self._lbl_rotate.pack(side="left", padx=(0, 3), pady=2)

        tk.Frame(ctrl2, bg=CLR["border"], width=1).pack(side="left", fill="y", padx=6, pady=4)

        self._btn(ctrl2, "−", self._zoom_out).pack(side="left", padx=(3, 1), pady=2)
        self._lbl_zoom = tk.Label(ctrl2, text="1.0×", font=FONT_SMALL,
                                  fg=CLR["text"], bg=CLR["btn"],
                                  width=5, cursor="hand2", relief="flat", padx=4, pady=5)
        self._lbl_zoom.pack(side="left", padx=1, pady=2)
        self._lbl_zoom.bind("<Button-1>", lambda e: self._zoom_reset())
        self._btn(ctrl2, "+", self._zoom_in).pack(side="left", padx=(1, 3), pady=2)

        # 우측 패널
        right = tk.Frame(self, bg=CLR["panel"], relief="flat")
        right.grid(row=1, column=1, sticky="nsew", padx=(0, 8), pady=8)
        right.columnconfigure(0, weight=1)
        right.rowconfigure(3, weight=1)

        tk.Label(right, text="INSPECTION RESULT", font=FONT_TITLE,
                 fg=CLR["accent"], bg=CLR["panel"]).grid(row=0, column=0,
                                                          sticky="w", padx=12, pady=(10, 4))
        self._lbl_verdict = tk.Label(right, text="—", font=("Consolas", 32, "bold"),
                                     fg=CLR["subtext"], bg=CLR["panel"])
        self._lbl_verdict.grid(row=1, column=0, pady=(0, 6))

        tk.Frame(right, bg=CLR["border"], height=1).grid(row=2, column=0,
                                                           sticky="ew", padx=12)

        self._txt = tk.Text(right, bg=CLR["bg"], fg=CLR["text"],
                            font=FONT_BODY, relief="flat", wrap="word",
                            state="disabled", padx=8, pady=8)
        self._txt.grid(row=3, column=0, sticky="nsew", padx=8, pady=8)
        sb = ttk.Scrollbar(right, command=self._txt.yview)
        sb.grid(row=3, column=1, sticky="ns", pady=8)
        self._txt.configure(yscrollcommand=sb.set)
        self._configure_text_tags()

        bot = tk.Frame(right, bg=CLR["panel"])
        bot.grid(row=4, column=0, columnspan=2, sticky="ew", padx=8, pady=(0, 8))
        self._btn(bot, "📊  Excel 내보내기", self._save_result).pack(side="left", padx=4)
        self._btn(bot, "📋  이력 보기",       self._show_history).pack(side="left", padx=4)
        self._btn(bot, "📌  기준값 저장",     self._set_ref_values,
                  bg=CLR["btn_on"]).pack(side="left", padx=4)

        self._statusbar = tk.Label(self, text="준비",
                                   font=FONT_SMALL, fg=CLR["subtext"],
                                   bg=CLR["border"], anchor="w", padx=8)
        self._statusbar.grid(row=2, column=0, columnspan=2, sticky="ew")

    # ── 위젯 헬퍼 ────────────────────────────────
    def _btn(self, parent, text, cmd, accent=False, bg=None):
        if bg is None:
            bg = CLR["accent"] if accent else CLR["btn"]
        fg = "#ffffff" if (accent or bg not in (CLR["btn"], None)) else CLR["text"]
        b  = tk.Button(parent, text=text, command=cmd, font=FONT_SMALL,
                       bg=bg, fg=fg, activebackground=CLR["btn_hover"],
                       activeforeground=CLR["text"], relief="flat",
                       padx=10, pady=5, cursor="hand2", bd=0)
        _bg = bg
        b.bind("<Enter>", lambda e: b.configure(bg=CLR["btn_hover"]))
        b.bind("<Leave>", lambda e: b.configure(bg=_bg))
        return b

    def _configure_text_tags(self):
        self._txt.tag_configure("ok",      foreground=CLR["ok"])
        self._txt.tag_configure("ng",      foreground=CLR["ng"])
        self._txt.tag_configure("warn",    foreground=CLR["warn"])
        self._txt.tag_configure("head",    foreground=CLR["accent"], font=FONT_TITLE)
        self._txt.tag_configure("subhead", foreground=CLR["subtext"], font=FONT_SMALL)
        self._txt.tag_configure("value",   foreground=CLR["text"])

    def _tick_clock(self):
        self._lbl_clock.configure(
            text=datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S"))
        self.after(1000, self._tick_clock)

    # ── 회전 / 확대 ──────────────────────────────
    _ROTATE_MAP = {
        0:   None,
        90:  cv2.ROTATE_90_CLOCKWISE,
        180: cv2.ROTATE_180,
        270: cv2.ROTATE_90_COUNTERCLOCKWISE,
    }

    def _apply_transform(self, bgr):
        code = self._ROTATE_MAP.get(self._rotate)
        if code is not None:
            bgr = cv2.rotate(bgr, code)
        return bgr

    def _rotate_cw(self):
        self._rotate = (self._rotate + 90) % 360
        self._lbl_rotate.configure(text=f"{self._rotate}°")
        self._refresh_static()

    def _rotate_ccw(self):
        self._rotate = (self._rotate - 90) % 360
        self._lbl_rotate.configure(text=f"{self._rotate}°")
        self._refresh_static()

    def _refresh_static(self):
        if not self.cam_running and not getattr(self, "_video_running", False):
            if self._static_bgr is not None:
                bgr = self._apply_transform(self._static_bgr.copy())
                self.current_bgr = bgr
                self._show_frame(bgr)

    def _redraw_current(self):
        """현재 프레임을 즉시 재렌더링 (줌/팬 변경 후 즉각 반영)."""
        if self.current_bgr is not None:
            self._show_frame(self.current_bgr, tmpl_preview=self.tmpl.loaded)

    def _zoom_in(self):
        if self._zoom < self._ZOOM_MAX:
            self._zoom = min(self._ZOOM_MAX, round(self._zoom + self._ZOOM_STEP, 1))
            self._lbl_zoom.configure(text=f"{self._zoom:.1f}×")
            self._redraw_current()

    def _zoom_out(self):
        if self._zoom > self._ZOOM_MIN:
            self._zoom = max(self._ZOOM_MIN, round(self._zoom - self._ZOOM_STEP, 1))
            self._lbl_zoom.configure(text=f"{self._zoom:.1f}×")
            self._redraw_current()

    def _zoom_reset(self):
        self._zoom  = 1.0
        self._pan_x = 0
        self._pan_y = 0
        self._lbl_zoom.configure(text="1.0×")
        self._redraw_current()

    def _on_mousewheel(self, event):
        if event.delta > 0:
            self._zoom_in()
        else:
            self._zoom_out()

    def _on_drag_start(self, event):
        if self._zoom > 1.0:
            self._drag_start = (event.x, event.y)
            self._canvas.configure(cursor="fleur")

    def _on_drag_move(self, event):
        if self._drag_start is None or self._zoom <= 1.0:
            return
        dx = event.x - self._drag_start[0]
        dy = event.y - self._drag_start[1]
        self._drag_start = (event.x, event.y)
        self._pan_x -= int(dx / self._zoom)
        self._pan_y -= int(dy / self._zoom)
        self._redraw_current()

    def _on_drag_end(self, event):
        self._drag_start = None
        self._canvas.configure(cursor="")

    # ── 해상도 변경 ──────────────────────────────
    def _on_res_change(self):
        if not self.cam_running:
            return
        rs = self._res_var.get().replace("×", "x")
        w, h = map(int, rs.split("x"))
        self._cam_res = (w, h)
        backend = getattr(self, "_current_cam_backend", cv2.CAP_DSHOW)
        self._stop_camera()
        self.after(400, lambda: self._start_camera(
            self._current_cam_idx, self._cam_fmt, (w, h), backend))

    # ── 카메라 ────────────────────────────────────
    def _start_camera(self, idx=0, fmt="YUY2", res=(640, 480), backend=cv2.CAP_DSHOW):
        self._current_cam_idx     = idx
        self._current_cam_backend = backend
        self._static_bgr = None
        self._btn_cam.configure(state="disabled")
        self._btn_stop.configure(state="disabled")
        bname = "MSMF" if backend == cv2.CAP_MSMF else "DSHOW"
        wait  = "  (~10초 소요)" if backend == cv2.CAP_MSMF else ""
        self._status(f"CAM {idx} [{bname}] 연결 중...{wait}")
        threading.Thread(target=self._open_camera_bg,
                         args=(idx, res, backend), daemon=True).start()

    @staticmethod
    def _get_dshow_devices():
        try:
            from pygrabber.dshow_graph import FilterGraph
            return FilterGraph().get_input_devices()
        except Exception:
            return []

    def _open_camera_bg(self, idx, res=(640, 480), backend=cv2.CAP_DSHOW):
        w, h   = res
        result = [None]

        def _do_open():
            try:
                c = cv2.VideoCapture(idx, backend)
                if c.isOpened():
                    if backend == cv2.CAP_MSMF:
                        c.set(cv2.CAP_PROP_FOURCC,
                              cv2.VideoWriter_fourcc(*'MJPG'))
                    c.set(cv2.CAP_PROP_FRAME_WIDTH,  w)
                    c.set(cv2.CAP_PROP_FRAME_HEIGHT, h)
                    result[0] = c
                else:
                    c.release()
            except Exception:
                pass

        t = threading.Thread(target=_do_open, daemon=True)
        t.start()
        timeout = 15.0 if backend == cv2.CAP_MSMF else 3.0
        t.join(timeout=timeout)

        cap = result[0]
        if cap is None and t.is_alive():
            self.after(0, lambda: self._status("연결 타임아웃 — 카메라 응답 없음"))
        self.after(0, lambda: self._on_camera_opened(cap, idx))

    def _on_camera_opened(self, cap, idx):
        self._btn_cam.configure(state="normal")
        if cap is None:
            self._btn_stop.configure(state="disabled")
            messagebox.showerror("오류", f"카메라 {idx}를 열 수 없습니다.")
            self._status(f"CAM {idx} 열기 실패")
            return
        self.cap           = cap
        self.cam_running   = True
        self._frame_count  = 0
        self._latest_frame = None   # 이전 세션 잔여 프레임 제거 + 크래시 방지
        self._btn_stop.configure(state="normal")
        rw = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        rh = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        self._status(f"CAM {idx} 스트리밍 중... {rw}×{rh}")
        threading.Thread(target=self._cam_reader, daemon=True).start()
        self.after(33, self._cam_display)
        self._start_tracking()

    def _stop_camera(self):
        self.cam_running   = False
        self._latest_frame = None
        self._stop_tracking()
        if self._after_id:
            self.after_cancel(self._after_id)
        if self.cap:
            self.cap.release()
            self.cap = None
        self._btn_cam.configure(state="normal")
        self._btn_stop.configure(state="disabled")
        if self._auto_inspect:
            self._toggle_auto_inspect()
        self._status("카메라 중지됨")

    # ── 라벨 추적 ─────────────────────────────────
    def _start_tracking(self):
        """ORB 특징점 매칭 추적 스레드 시작."""
        if self._tracking or not self.tmpl.has_tracking:
            return
        self._tracking = True
        self._track_H  = None
        self._track_ok = False
        threading.Thread(target=self._track_loop, daemon=True).start()

    def _stop_tracking(self):
        self._tracking = False
        self._track_H  = None
        self._track_ok = False

    def _track_loop(self):
        """백그라운드에서 4fps로 호모그래피 갱신."""
        import time as _t
        while self._tracking:
            frame = self.current_bgr
            if frame is not None and self.tmpl.has_tracking:
                try:
                    H = self.tmpl.match_frame(frame)
                    self._track_H  = H
                    self._track_ok = H is not None
                except Exception:
                    self._track_H  = None
                    self._track_ok = False
            _t.sleep(0.25)
        self._tracking = False

    def _cam_reader(self):
        import time as _time
        fail      = 0
        warmup    = 15
        _interval = 1 / 20   # 20fps 상한
        _last_t   = -999.0
        try:
            while self.cam_running and self.cap is not None:
                ret, frame = self.cap.read()
                if not ret or frame is None:
                    fail += 1
                    if fail >= 30:
                        self.after(0, lambda: self._status("프레임 없음 — 재연결 필요"))
                    _time.sleep(0.01)
                    continue
                fail = 0
                if warmup > 0:
                    warmup -= 1
                    continue
                mean = frame.mean()
                if mean < 2.0 or mean > 253.0:
                    continue
                now = _time.monotonic()
                if now - _last_t < _interval:
                    continue          # 이번 프레임 스킵, 다음 cap.read()로
                _last_t = now
                self._latest_frame = frame
                self._frame_count  = getattr(self, "_frame_count", 0) + 1
        except Exception as e:
            self.after(0, lambda: self._status(f"카메라 읽기 오류: {e}"))

    def _cam_display(self):
        if not self.cam_running:
            return
        frame = self._latest_frame
        if frame is not None:
            frame = self._apply_transform(frame)
            self.current_bgr   = frame   # 검사용 풀 해상도 보존
            self._latest_frame = None
            try:
                # 디스플레이용 다운스케일 (캔버스 크기에 맞게 미리 축소)
                fh, fw = frame.shape[:2]
                cw = self._canvas.winfo_width()  or self.PREVIEW_W
                ch = self._canvas.winfo_height() or self.PREVIEW_H
                ds = min(cw / fw, ch / fh)
                if ds < 0.9:
                    disp = cv2.resize(frame, (int(fw * ds), int(fh * ds)),
                                      interpolation=cv2.INTER_AREA)
                    disp_H = None
                    if self._track_H is not None:
                        S = np.array([[ds, 0, 0], [0, ds, 0], [0, 0, 1]], np.float64)
                        disp_H = S @ self._track_H
                else:
                    disp, disp_H = frame, self._track_H
                self._show_frame(disp, tmpl_preview=self.tmpl.loaded, disp_H=disp_H)
                fc = getattr(self, "_frame_count", 0)
                zoom_str = f"  줌:{self._zoom:.1f}×" if self._zoom > 1.0 else ""
                self._status(f"스트리밍 중... {fw}×{fh}  밝기:{int(frame.mean())}  프레임:{fc}{zoom_str}")
            except Exception as e:
                self._status(f"표시 오류: {e}")
        self._after_id = self.after(50, self._cam_display)

    # ── 파일 열기 ────────────────────────────────
    def _open_file(self):
        path = filedialog.askopenfilename(
            title="파일 선택",
            filetypes=[
                ("이미지/동영상", "*.jpg *.jpeg *.png *.bmp *.tiff *.tif *.mp4 *.avi *.mov *.mkv"),
                ("이미지",       "*.jpg *.jpeg *.png *.bmp *.tiff *.tif"),
                ("동영상",       "*.mp4 *.avi *.mov *.mkv"),
                ("모든 파일",    "*.*"),
            ]
        )
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        if ext in (".mp4", ".avi", ".mov", ".mkv"):
            self._open_video(path)
        else:
            bgr = cv2.imdecode(np.fromfile(path, dtype=np.uint8), cv2.IMREAD_COLOR)
            if bgr is None:
                messagebox.showerror("오류", "이미지를 읽을 수 없습니다.")
                return
            self._stop_camera()
            self._stop_video()
            self._static_bgr = bgr.copy()
            bgr = self._apply_transform(bgr)
            self.current_bgr = bgr
            self._show_frame(bgr)
            self._status(f"이미지 로드: {os.path.basename(path)}")

    def _open_video(self, path):
        vcap = cv2.VideoCapture(path)
        if not vcap.isOpened():
            messagebox.showerror("오류", "동영상 파일을 열 수 없습니다.")
            return
        self._stop_camera()
        self._stop_video()
        self._static_bgr    = None
        self._vcap          = vcap
        self._video_running = True
        total = int(vcap.get(cv2.CAP_PROP_FRAME_COUNT))
        fps   = vcap.get(cv2.CAP_PROP_FPS) or 30
        self._video_delay   = max(15, int(1000 / fps))
        self._status(f"동영상: {os.path.basename(path)}  ({total}프레임, {fps:.1f}fps)")
        self._video_loop()

    def _stop_video(self):
        self._video_running = False
        if getattr(self, "_video_after_id", None):
            self.after_cancel(self._video_after_id)
            self._video_after_id = None
        if getattr(self, "_vcap", None):
            self._vcap.release()
            self._vcap = None

    def _video_loop(self):
        if not getattr(self, "_video_running", False) or self._vcap is None:
            return
        ret, frame = self._vcap.read()
        if ret and frame is not None:
            frame = self._apply_transform(frame)
            self.current_bgr = frame
            self._show_frame(frame, tmpl_preview=self.tmpl.loaded)
            pos   = int(self._vcap.get(cv2.CAP_PROP_POS_FRAMES))
            total = int(self._vcap.get(cv2.CAP_PROP_FRAME_COUNT))
            self._status(f"동영상 재생 중... ({pos}/{total}프레임)")
        else:
            self._vcap.set(cv2.CAP_PROP_POS_FRAMES, 0)
        self._video_after_id = self.after(self._video_delay, self._video_loop)

    # ── 프레임 표시 ──────────────────────────────
    def _show_frame(self, bgr, result_regions=None, tmpl_preview=False, disp_H=None):
        if bgr.ndim == 2:
            bgr = cv2.cvtColor(bgr, cv2.COLOR_GRAY2BGR)
        elif bgr.ndim == 3 and bgr.shape[2] == 4:
            bgr = bgr[:, :, :3]

        rgb = cv2.cvtColor(bgr.copy(), cv2.COLOR_BGR2RGB)

        if self.tmpl.loaded:
            ih, iw = bgr.shape[:2]
            H = disp_H if disp_H is not None else self._track_H
            regs = self.tmpl.data.get("regions", [])

            if H is None:
                tw, th = self.tmpl.data.get("img_size", [iw, ih])
                sx = iw / tw if tw else 1.0
                sy = ih / th if th else 1.0

            def _draw_region(pts_or_rect, color, thickness, label=""):
                """4점(기울기) 또는 [x,y,w,h](축정렬) 모두 처리."""
                if isinstance(pts_or_rect, np.ndarray):
                    # 4개 꼭짓점 → polylines
                    pts_i = pts_or_rect.astype(np.int32).reshape((-1, 1, 2))
                    cv2.polylines(rgb, [pts_i], isClosed=True,
                                  color=color, thickness=thickness)
                    lx, ly = int(pts_or_rect[0, 0]), int(pts_or_rect[0, 1])
                else:
                    rx, ry, rw, rh = pts_or_rect
                    cv2.rectangle(rgb, (rx, ry), (rx+rw, ry+rh), color, thickness)
                    lx, ly = rx, ry
                if label:
                    cv2.putText(rgb, label, (lx + 2, max(12, ly - 3)),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.4, color, 1)

            def _get_pts(reg):
                """H가 있으면 기울어진 4점, 없으면 축정렬 rect 반환."""
                if H is not None:
                    return TemplateManager.transform_corners(H, reg["rect"])
                r = reg["rect"]
                return [int(r[0]*sx), int(r[1]*sy), int(r[2]*sx), int(r[3]*sy)]

            if result_regions:
                # 검사 결과: OK=녹색 / NG=빨간색
                for i, res in enumerate(result_regions):
                    if i >= len(regs):
                        break
                    color = (0, 230, 100) if res["ok"] else (255, 61, 61)
                    _draw_region(_get_pts(regs[i]), color, 2, res["label"])
            elif tmpl_preview:
                # 템플릿 미리보기: 유형별 색상 윤곽 + 라벨
                for reg in regs:
                    color = _hex_to_rgb(REGION_COL.get(reg["type"], "#80c8ff"))
                    _draw_region(_get_pts(reg), color, 1, reg.get("label", reg["type"]))

            # 추적 상태 표시기 (우하단)
            if self.tmpl.has_tracking:
                ok_c  = (0, 210, 80)   if self._track_ok else (200, 160,  0)
                label = "TRACKING"     if self._track_ok else "SEARCHING..."
                tw_px = len(label) * 9 + 12
                cv2.rectangle(rgb, (iw - tw_px - 4, ih - 22), (iw - 4, ih - 4), ok_c, -1)
                cv2.putText(rgb, label, (iw - tw_px, ih - 8),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.42, (0, 0, 0), 1, cv2.LINE_AA)

        if self._zoom > 1.0:
            fh, fw = rgb.shape[:2]
            zh, zw = int(fh / self._zoom), int(fw / self._zoom)
            x0 = max(0, min(fw - zw, (fw - zw) // 2 + self._pan_x))
            y0 = max(0, min(fh - zh, (fh - zh) // 2 + self._pan_y))
            rgb = rgb[y0:y0 + zh, x0:x0 + zw]

        cw = self._canvas.winfo_width()
        ch = self._canvas.winfo_height()
        if cw <= 1:  cw = self.PREVIEW_W
        if ch <= 1:  ch = self.PREVIEW_H
        fh, fw = rgb.shape[:2]
        scale  = min(cw / fw, ch / fh)
        nw, nh = int(fw * scale), int(fh * scale)
        rgb    = cv2.resize(rgb, (nw, nh), interpolation=cv2.INTER_AREA)

        img = ImageTk.PhotoImage(Image.fromarray(rgb))
        self._photo_ref = img
        self._canvas.delete("hint")
        self._canvas.delete("frame")
        self._canvas.create_image(cw // 2, ch // 2, image=img, anchor="center", tags="frame")
        self._canvas.update_idletasks()

    # ── 검사 ─────────────────────────────────────
    def _inspect(self):
        if self.current_bgr is None:
            messagebox.showwarning("경고", "카메라를 시작하거나 이미지 파일을 열어주세요.")
            return
        if not self.tmpl.loaded:
            messagebox.showwarning("경고", "템플릿을 먼저 불러오거나 생성해주세요.")
            return

        self._btn_snap.configure(state="disabled", text="분석 중...")
        self._status("검사 진행 중...")
        frame = self.current_bgr.copy()

        H = self._track_H  # 스냅샷 (스레드 안전)

        def worker():
            try:
                region_results = self.tmpl.inspect(frame, H=H)
                self.after(0, lambda: self._on_result(region_results, frame))
            except Exception:
                import traceback
                tb = traceback.format_exc()
                self.after(0, lambda: self._on_error(tb))

        threading.Thread(target=worker, daemon=True).start()

    def _on_error(self, msg):
        self._btn_snap.configure(state="normal", text="🔍  검사 실행")
        self._status("❌ 오류 발생")
        self._lbl_verdict.configure(text="[ ERR ]", fg=CLR["warn"])
        self._txt.configure(state="normal")
        self._txt.delete("1.0", "end")
        self._txt.insert("end", "오류 발생:\n\n", "head")
        self._txt.insert("end", msg, "ng")
        self._txt.configure(state="disabled")

    def _on_result(self, region_results, frame):
        try:
            self._btn_snap.configure(state="normal", text="🔍  검사 실행")
            verdict = "OK" if region_results and all(r["ok"] for r in region_results) else "NG"
            entry = {
                "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "verdict":   verdict,
                "template":  self.tmpl.name,
                "regions":   region_results,
            }
            self.history.append(entry)
            self._render_result(entry)
            self._show_frame(frame, result_regions=region_results)
            self._lbl_verdict.configure(
                text=f"[ {verdict} ]",
                fg=CLR["ok"] if verdict == "OK" else CLR["ng"])
            self._append_csv_log(entry)
            self._status(
                f"검사 완료 — {entry['timestamp']}  →  {verdict}"
                f"  |  로그: {os.path.basename(self._log_path)}")
        except Exception:
            import traceback
            self._on_error(traceback.format_exc())

    # ── 결과 렌더링 ──────────────────────────────
    def _render_result(self, entry):
        self._txt.configure(state="normal")
        self._txt.delete("1.0", "end")

        def w(text, tag="value"):
            self._txt.insert("end", text, tag)

        w(f"■ 검사 결과  [{entry['template']}]\n", "head")
        w(f"  {entry['timestamp']}\n\n", "subhead")

        for r in entry.get("regions", []):
            tag = "ok" if r["ok"] else "ng"
            w(f"  {'[OK]' if r['ok'] else '[NG]'} ", tag)
            w(f"{r['label']:<12}", "subhead")
            val = r["value"][:40] if r["value"] else "(없음)"
            w(f"{val}", "value")
            if r["reason"]:
                w(f"  ← {r['reason']}", "warn")
            w("\n")

        self._txt.configure(state="disabled")

    # ── CSV 로그 ─────────────────────────────────
    def _ensure_csv_header(self):
        if not os.path.exists(self._log_path):
            with open(self._log_path, "w", newline="", encoding="utf-8-sig") as f:
                csv.writer(f).writerow(self._CSV_COLS)

    def _append_csv_log(self, entry):
        ts      = entry["timestamp"]
        verdict = entry["verdict"]
        tmpl    = entry["template"]
        with open(self._log_path, "a", newline="", encoding="utf-8-sig") as f:
            wr = csv.writer(f)
            for r in entry.get("regions", []):
                wr.writerow([ts, verdict, tmpl,
                             r["label"], r["type"],
                             "OK" if r["ok"] else "NG",
                             r["value"], r["reason"]])

    # ── 기준값 설정 ──────────────────────────────
    def _set_ref_values(self):
        """마지막 검사 결과를 템플릿 기준값으로 저장."""
        if not self.history:
            messagebox.showinfo("알림", "검사 결과가 없습니다.")
            return
        if not self.tmpl.loaded:
            messagebox.showinfo("알림", "로드된 템플릿이 없습니다.")
            return

        last    = self.history[-1]
        regions = self.tmpl.data.get("regions", [])
        reg_map = {r.get("label", ""): i for i, r in enumerate(regions)}

        d = tk.Toplevel(self)
        d.title("기준값 설정")
        d.configure(bg=CLR["bg"])
        d.resizable(False, False)
        d.grab_set()

        tk.Label(d, text="검사 결과를 기준값으로 저장합니다.",
                 font=FONT_BODY, fg=CLR["text"], bg=CLR["bg"]).pack(padx=16, pady=(14, 2))
        tk.Label(d,
                 text="체크된 항목은 다음 검사부터 기준값과 비교합니다.  비어 있으면 존재 여부만 확인.",
                 font=FONT_SMALL, fg=CLR["subtext"], bg=CLR["bg"]).pack(padx=16, pady=(0, 10))

        frm = tk.Frame(d, bg=CLR["panel"], bd=0)
        frm.pack(fill="both", padx=16, pady=4)

        checks = []
        for r in last.get("regions", []):
            typ       = r.get("type", "")
            cur_val   = r.get("value", "")
            # 이미 설정된 기준값이 있으면 그 값을, 없으면 현재 인식 값 사용
            reg_idx   = reg_map.get(r["label"])
            saved_ref = ""
            if reg_idx is not None:
                saved_ref = regions[reg_idx].get("text", "")
            init_val  = saved_ref if saved_ref else cur_val

            var       = tk.BooleanVar(value=bool(init_val))
            entry_var = tk.StringVar(value=init_val)

            row = tk.Frame(frm, bg=CLR["panel"])
            row.pack(fill="x", padx=8, pady=3)

            tk.Checkbutton(row, variable=var, bg=CLR["panel"],
                           fg=CLR["text"], selectcolor=CLR["btn"],
                           activebackground=CLR["panel"]).pack(side="left")
            tk.Label(row, text=f"{r['label']}", font=FONT_BODY,
                     fg=CLR["subtext"], bg=CLR["panel"], width=16, anchor="w").pack(side="left")

            if typ == "barcode":
                # 바코드: 정확 일치 기준값
                e = tk.Entry(row, textvariable=entry_var, font=FONT_BODY,
                             bg=CLR["btn"], fg=CLR["text"],
                             insertbackground=CLR["text"], relief="flat", width=30)
                e.pack(side="left", padx=(4, 0))
                tk.Label(row, text=" (바코드 정확 일치)", font=FONT_SMALL,
                         fg=CLR["subtext"], bg=CLR["panel"]).pack(side="left", padx=4)
            else:
                # 텍스트/이미지 구역: OCR 포함 여부 비교
                e = tk.Entry(row, textvariable=entry_var, font=FONT_BODY,
                             bg=CLR["btn"], fg=CLR["text"],
                             insertbackground=CLR["text"], relief="flat", width=30)
                e.pack(side="left", padx=(4, 0))
                note = " (인식값 포함 여부)" if cur_val else " (비어 있음 — 존재 여부만 확인)"
                tk.Label(row, text=note, font=FONT_SMALL,
                         fg=CLR["subtext"], bg=CLR["panel"]).pack(side="left", padx=4)

            # 현재 인식 결과 표시
            detected = f"인식: {cur_val[:28]}" if cur_val else "인식: (없음)"
            tk.Label(row, text=detected, font=FONT_SMALL,
                     fg=CLR["ok"] if r.get("ok") else CLR["ng"],
                     bg=CLR["panel"]).pack(side="right", padx=8)

            checks.append((r["label"], var, entry_var))

        def _apply():
            changed = False
            for label, var, entry_var in checks:
                idx = reg_map.get(label)
                if idx is None:
                    continue
                regions[idx]["text"] = entry_var.get().strip() if var.get() else ""
                changed = True
            if changed and self.tmpl.path:
                try:
                    with open(self.tmpl.path, "w", encoding="utf-8") as f:
                        json.dump(self.tmpl.data, f, ensure_ascii=False, indent=2)
                    self._status(f"기준값 저장 완료 → {os.path.basename(self.tmpl.path)}")
                except Exception as e:
                    messagebox.showerror("오류", f"저장 실패: {e}", parent=d)
            d.destroy()

        def _clear_all():
            for _, var, entry_var in checks:
                var.set(False)
                entry_var.set("")

        bot_d = tk.Frame(d, bg=CLR["bg"])
        bot_d.pack(fill="x", padx=16, pady=12)
        self._btn(bot_d, "저장", _apply, accent=True).pack(side="left", padx=(0, 8))
        self._btn(bot_d, "전체 초기화", _clear_all).pack(side="left", padx=(0, 8))
        self._btn(bot_d, "닫기", d.destroy).pack(side="left")

        d.update_idletasks()
        x = self.winfo_x() + max(0, (self.winfo_width()  - d.winfo_reqwidth())  // 2)
        y = self.winfo_y() + max(0, (self.winfo_height() - d.winfo_reqheight()) // 2)
        d.geometry(f"+{x}+{y}")

    # ── 저장 / 내보내기 ───────────────────────────
    def _save_result(self):
        if not self.history:
            messagebox.showinfo("알림", "저장할 결과가 없습니다.")
            return
        if OPENPYXL_OK:
            default_ext = ".xlsx"
            ftypes = [("Excel 통합 문서", "*.xlsx"), ("CSV", "*.csv"), ("모든 파일", "*.*")]
        else:
            default_ext = ".csv"
            ftypes = [("CSV", "*.csv"), ("모든 파일", "*.*")]
        path = filedialog.asksaveasfilename(
            defaultextension=default_ext, filetypes=ftypes,
            initialfile=f"label_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}{default_ext}"
        )
        if not path:
            return
        if path.lower().endswith(".csv"):
            self._export_csv(path)
        else:
            self._export_xlsx(path)

    def _export_csv(self, path):
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            wr = csv.writer(f)
            wr.writerow(self._CSV_COLS)
            for entry in self.history:
                for r in entry.get("regions", []):
                    wr.writerow([entry["timestamp"], entry["verdict"], entry["template"],
                                 r["label"], r["type"],
                                 "OK" if r["ok"] else "NG",
                                 r["value"], r["reason"]])
        self._status(f"CSV 저장 완료: {os.path.basename(path)}")

    def _export_xlsx(self, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Label Log"
        hdr_fill = PatternFill("solid", fgColor="1A3A5C")
        hdr_font = Font(bold=True, color="FFFFFF", name="Consolas", size=10)
        ok_fill  = PatternFill("solid", fgColor="C8F0D4")
        ng_fill  = PatternFill("solid", fgColor="FFD0D0")
        center   = Alignment(horizontal="center", vertical="center")

        for ci, name in enumerate(self._CSV_COLS, 1):
            c = ws.cell(row=1, column=ci, value=name)
            c.fill = hdr_fill;  c.font = hdr_font;  c.alignment = center

        row_i = 2
        for entry in self.history:
            for r in entry.get("regions", []):
                ok_str = "OK" if r["ok"] else "NG"
                vals = [entry["timestamp"], entry["verdict"], entry["template"],
                        r["label"], r["type"], ok_str, r["value"], r["reason"]]
                for ci, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_i, column=ci, value=val)
                    if ci in (2, 6):
                        cell.fill = ok_fill if val == "OK" else ng_fill
                        cell.font = Font(bold=True, name="Consolas", size=10)
                        cell.alignment = center
                row_i += 1

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        ws.freeze_panes = "A2"
        wb.save(path)
        self._status(f"Excel 저장 완료: {os.path.basename(path)}")

    # ── 템플릿 마법사 ────────────────────────────
    def _template_wizard(self, initial_data=None):
        """드래그로 구역 지정 → 템플릿 JSON 저장. initial_data 있으면 편집 모드."""

        _COL = dict(REGION_COL)   # 모듈 상수에서 복사 (custom type 추가 가능)
        _LBL = {
            "desc":    "제품 설명",
            "pn":      "P/N",
            "sn":      "S/N",
            "ocr":     "OCR 기타",
            "barcode": "바코드",
            "ci":      "회사 CI",
            "class1":  "Class 1 Laser",
            "certi":   "Certi. Mark",
        }
        _EXTRA_PAL = REGION_COL_EXTRA

        # 기존 템플릿의 custom type 사전 등록
        if initial_data:
            _ei = 0
            for _reg in initial_data.get("regions", []):
                _t = _reg.get("type", "")
                if _t.startswith("cust_") and _t not in _COL:
                    _COL[_t] = _EXTRA_PAL[_ei % len(_EXTRA_PAL)]
                    _LBL[_t] = _reg.get("label", _t)
                    _ei += 1

        def _open_editor(bgr_orig):
            ih, iw = bgr_orig.shape[:2]
            bgr = bgr_orig.copy()
            is_edit = initial_data is not None

            win = tk.Toplevel(self)
            win.title("템플릿 편집 — 구역 설정" if is_edit else "템플릿 생성 — 구역 설정")
            win.configure(bg=CLR["bg"])
            win.geometry("1100x720")
            win.resizable(True, True)
            win.grab_set()

            # 기존 구역을 현재 이미지 크기에 맞게 스케일해 pre-populate
            regions = []
            if initial_data:
                tw, th = initial_data.get("img_size", [iw, ih])
                sx = iw / tw if tw else 1.0
                sy = ih / th if th else 1.0
                for reg in initial_data.get("regions", []):
                    rx, ry, rw, rh = reg["rect"]
                    regions.append({
                        "label": reg.get("label", ""),
                        "type":  reg.get("type", "ocr"),
                        "rect":  (int(rx * sx), int(ry * sy),
                                  max(4, int(rw * sx)), max(4, int(rh * sy))),
                        "text":  reg.get("text", ""),
                        "cids":  [],
                    })
            zoom   = [1.0]
            pan    = [0, 0]
            ri     = {"ox": 0, "oy": 0, "sc": 1.0}
            tk_ref = [None]

            toolbar   = tk.Frame(win, bg=CLR["bg"])
            toolbar.pack(fill="x", padx=8, pady=(6, 3))
            body      = tk.Frame(win, bg=CLR["bg"])
            body.pack(fill="both", expand=True, padx=8, pady=(0, 8))
            canvas_fr = tk.Frame(body, bg="#0d0f13")
            canvas_fr.pack(side="left", fill="both", expand=True)
            right_fr  = tk.Frame(body, bg=CLR["panel"], width=265)
            right_fr.pack(side="right", fill="y", padx=(6, 0))
            right_fr.pack_propagate(False)

            canvas = tk.Canvas(canvas_fr, bg="#0d0f13", cursor="crosshair",
                               highlightthickness=0)
            canvas.pack(fill="both", expand=True)

            # ── 영역 박스 그리기 ──────────────────
            def _draw_r(r):
                for cid in r.get("cids", []):
                    canvas.delete(cid)
                x, y, w, h = r["rect"]
                ox_ = ri["ox"]; oy_ = ri["oy"]; sc_ = ri["sc"]
                cx  = ox_ + int(x * sc_)
                cy  = oy_ + int(y * sc_)
                cw2 = max(4, int(w * sc_))
                ch2 = max(4, int(h * sc_))
                col = _COL.get(r["type"], "#aaaaaa")
                ids = [
                    canvas.create_rectangle(cx, cy, cx + cw2, cy + ch2,
                                            outline=col, width=2,
                                            fill=col, stipple="gray12"),
                ]
                lbl  = _LBL.get(r["type"], r.get("type", "기타"))
                disp = (r.get("label") or r.get("text", ""))[:18]
                ids.append(canvas.create_text(
                    cx + 3, cy - 2, anchor="sw",
                    text=f"[{lbl}] {disp}",
                    font=("Consolas", 8), fill=col,
                ))
                r["cids"] = ids

            # ── 렌더링 ───────────────────────────
            def _render(event=None):
                cw_ = canvas.winfo_width()
                ch_ = canvas.winfo_height()
                if cw_ < 10 or ch_ < 10:
                    return
                fit = min(cw_ / iw, ch_ / ih)
                sc_ = fit * zoom[0]
                dw_ = max(1, int(iw * sc_))
                dh_ = max(1, int(ih * sc_))
                ox_ = (cw_ - dw_) // 2 + pan[0]
                oy_ = (ch_ - dh_) // 2 + pan[1]
                ri["ox"] = ox_;  ri["oy"] = oy_;  ri["sc"] = sc_
                rgb = cv2.cvtColor(
                    cv2.resize(bgr, (dw_, dh_), interpolation=cv2.INTER_LINEAR),
                    cv2.COLOR_BGR2RGB)
                tk_img = ImageTk.PhotoImage(Image.fromarray(rgb))
                tk_ref[0] = tk_img
                canvas.delete("all")
                canvas.create_image(ox_, oy_, anchor="nw", image=tk_img)
                for r in regions:
                    _draw_r(r)

            canvas.bind("<Configure>", _render)

            def _on_scroll(event):
                factor = 1.15 if event.delta > 0 else (1 / 1.15)
                zoom[0] = max(0.2, min(10.0, zoom[0] * factor))
                _render()

            canvas.bind("<MouseWheel>", _on_scroll)

            _pan_s = {"xy": None, "base": None}

            def _pan_start(event):
                _pan_s["xy"]   = (event.x, event.y)
                _pan_s["base"] = list(pan)
                canvas.configure(cursor="fleur")

            def _pan_move(event):
                if _pan_s["xy"] is None:
                    return
                pan[0] = _pan_s["base"][0] + event.x - _pan_s["xy"][0]
                pan[1] = _pan_s["base"][1] + event.y - _pan_s["xy"][1]
                _render()

            def _pan_end(event):
                _pan_s["xy"] = None
                canvas.configure(cursor="crosshair")

            canvas.bind("<ButtonPress-2>",   _pan_start)
            canvas.bind("<B2-Motion>",       _pan_move)
            canvas.bind("<ButtonRelease-2>", _pan_end)
            canvas.bind("<ButtonPress-3>",   _pan_start)
            canvas.bind("<B3-Motion>",       _pan_move)
            canvas.bind("<ButtonRelease-3>", _pan_end)

            # ── 분류 팝업 ────────────────────────
            sel_type = tk.StringVar(value="pn")

            def _classify(r):
                d = tk.Toplevel(win)
                d.title("구역 설정")
                d.configure(bg=CLR["bg"])
                d.geometry("290x470")
                d.grab_set()
                d.transient(win)
                d.resizable(False, False)

                # 확인/삭제/취소 버튼을 하단에 먼저 고정 → 유형 목록이 길어도 항상 보임
                bot2 = tk.Frame(d, bg=CLR["bg"])
                bot2.pack(side="bottom", fill="x", padx=12, pady=10)

                tk.Label(d, text="표시 이름 (label)", font=FONT_SMALL,
                         fg=CLR["subtext"], bg=CLR["bg"]).pack(anchor="w", padx=12, pady=(10, 2))
                e_label = tk.Entry(d, font=FONT_BODY, bg=CLR["btn"], fg=CLR["text"],
                                   insertbackground=CLR["text"], relief="flat")
                e_label.insert(0, r.get("label") or _LBL.get(r.get("type",""), ""))
                e_label.pack(fill="x", padx=12, pady=(0, 6))

                tk.Label(d, text="기대 텍스트 (빈칸 = 존재 여부만)", font=FONT_SMALL,
                         fg=CLR["subtext"], bg=CLR["bg"]).pack(anchor="w", padx=12, pady=(2, 2))
                e_txt = tk.Entry(d, font=FONT_BODY, bg=CLR["btn"], fg=CLR["text"],
                                 insertbackground=CLR["text"], relief="flat")
                e_txt.insert(0, r.get("text", ""))
                e_txt.pack(fill="x", padx=12, pady=(0, 8))

                tk.Label(d, text="유형 선택", font=FONT_SMALL,
                         fg=CLR["subtext"], bg=CLR["bg"]).pack(anchor="w", padx=12)
                typ_var = tk.StringVar(value=r.get("type") or sel_type.get())
                for val, lbl_text in _LBL.items():
                    row2 = tk.Frame(d, bg=CLR["bg"])
                    row2.pack(anchor="w", padx=16, pady=2)
                    tk.Frame(row2, bg=_COL[val], width=10, height=10).pack(
                        side="left", padx=(0, 6))
                    tk.Radiobutton(row2, text=lbl_text, variable=typ_var, value=val,
                                   font=FONT_SMALL, fg=CLR["text"], bg=CLR["bg"],
                                   activebackground=CLR["bg"],
                                   selectcolor=CLR["btn"]).pack(side="left")

                def _delete():
                    for cid in r.get("cids", []):
                        canvas.delete(cid)
                    if r in regions:
                        regions.remove(r)
                    d.destroy()
                    _refresh_list()
                    _render()

                def _ok():
                    r["type"]  = typ_var.get()
                    r["text"]  = e_txt.get().strip()
                    r["label"] = e_label.get().strip() or _LBL.get(r["type"], r["type"])
                    sel_type.set(r["type"])
                    _draw_r(r)
                    _refresh_list()
                    d.destroy()

                self._btn(bot2, "확인", _ok, accent=True).pack(side="left", padx=(0, 4))
                self._btn(bot2, "삭제", _delete).pack(side="left", padx=(0, 4))
                self._btn(bot2, "취소", d.destroy).pack(side="left")

            # ── 드래그 구역 추가 ─────────────────
            _ds = [None]
            _dt = [None]

            def _img_xy(cx, cy):
                return (cx - ri["ox"]) / ri["sc"], (cy - ri["oy"]) / ri["sc"]

            def _hit_region(cx, cy):
                ix, iy = _img_xy(cx, cy)
                for r in reversed(regions):
                    x, y, w, h = r["rect"]
                    if x <= ix <= x + w and y <= iy <= y + h:
                        return r
                return None

            def _ldown(event):
                _ds[0] = (event.x, event.y)

            def _lmove(event):
                if _ds[0] is None:
                    return
                if _dt[0]:
                    canvas.delete(_dt[0])
                x0, y0 = _ds[0]
                col = _COL.get(sel_type.get(), "#ff4444")
                _dt[0] = canvas.create_rectangle(
                    x0, y0, event.x, event.y,
                    outline=col, width=2, dash=(4, 3), fill="")

            def _lup(event):
                try:
                    start = _ds[0];  _ds[0] = None
                    if _dt[0]:
                        canvas.delete(_dt[0]);  _dt[0] = None
                    if start is None:
                        return
                    x0, y0 = start
                    x1, y1 = event.x, event.y
                    if abs(x1 - x0) < 5 and abs(y1 - y0) < 5:
                        hit = _hit_region(x0, y0)
                        if hit:
                            _classify(hit)
                        return
                    ix0, iy0 = _img_xy(min(x0, x1), min(y0, y1))
                    ix1, iy1 = _img_xy(max(x0, x1), max(y0, y1))
                    ix0 = max(0, int(ix0));  iy0 = max(0, int(iy0))
                    ix1 = min(iw, int(ix1)); iy1 = min(ih, int(iy1))
                    if ix1 <= ix0 or iy1 <= iy0:
                        messagebox.showwarning("알림", "이미지 범위 밖입니다.", parent=win)
                        return
                    typ = sel_type.get()
                    r   = {"label": _LBL.get(typ, typ), "text": "",
                           "rect": (ix0, iy0, ix1 - ix0, iy1 - iy0),
                           "type": typ, "cids": []}
                    if typ == "barcode" and PYZBAR_OK:
                        try:
                            codes = pyzbar.decode(bgr[iy0:iy1, ix0:ix1])
                            if codes:
                                r["text"] = codes[0].data.decode("utf-8", errors="replace")
                        except Exception:
                            pass
                    regions.append(r)
                    _draw_r(r)
                    _refresh_list()
                    _classify(r)
                except Exception:
                    import traceback
                    messagebox.showerror("오류", traceback.format_exc(), parent=win)

            canvas.bind("<ButtonPress-1>",   _ldown)
            canvas.bind("<B1-Motion>",        _lmove)
            canvas.bind("<ButtonRelease-1>",  _lup)

            # ── 오른쪽 패널 ──────────────────────
            tk.Label(right_fr, text="템플릿 이름", font=FONT_SMALL,
                     fg=CLR["subtext"], bg=CLR["panel"]).pack(
                anchor="w", padx=10, pady=(10, 2))
            e_name = tk.Entry(right_fr, font=FONT_BODY, bg=CLR["btn"], fg=CLR["text"],
                              insertbackground=CLR["text"], relief="flat")
            e_name.insert(0, initial_data.get("template_name", "") if is_edit else "")
            e_name.pack(fill="x", padx=10, pady=(0, 8))

            tk.Frame(right_fr, bg=CLR["border"], height=1).pack(fill="x", padx=10)

            tk.Label(right_fr, text="드래그 유형", font=FONT_SMALL,
                     fg=CLR["subtext"], bg=CLR["panel"]).pack(
                anchor="w", padx=10, pady=(8, 2))

            type_radio_fr = tk.Frame(right_fr, bg=CLR["panel"])
            type_radio_fr.pack(anchor="w", fill="x")

            _extra_idx = [sum(1 for k in _COL if k.startswith("cust_"))]

            def _rebuild_type_radios():
                for w_ in type_radio_fr.winfo_children():
                    w_.destroy()
                for val, lbl_text in _LBL.items():
                    row2 = tk.Frame(type_radio_fr, bg=CLR["panel"])
                    row2.pack(anchor="w", padx=10, pady=2)
                    tk.Frame(row2, bg=_COL[val], width=10, height=10).pack(
                        side="left", padx=(0, 5))
                    tk.Radiobutton(row2, text=lbl_text, variable=sel_type, value=val,
                                   font=FONT_SMALL, fg=CLR["text"], bg=CLR["panel"],
                                   activebackground=CLR["panel"],
                                   selectcolor=CLR["btn"]).pack(side="left")

            _rebuild_type_radios()

            tk.Frame(right_fr, bg=CLR["border"], height=1).pack(fill="x", padx=10, pady=(4, 0))
            add_row = tk.Frame(right_fr, bg=CLR["panel"])
            add_row.pack(fill="x", padx=10, pady=(4, 2))
            e_custom = tk.Entry(add_row, font=FONT_SMALL, bg=CLR["btn"], fg=CLR["text"],
                                insertbackground=CLR["text"], relief="flat", width=10)
            e_custom.pack(side="left", fill="x", expand=True, padx=(0, 4))
            e_custom.insert(0, "유형 이름")
            e_custom.bind("<FocusIn>", lambda e: e_custom.delete(0, "end")
                          if e_custom.get() == "유형 이름" else None)

            def _add_custom_type(event=None):
                name = e_custom.get().strip()
                if not name or name == "유형 이름":
                    return
                key = f"cust_{len([k for k in _COL if k.startswith('cust_')])}"
                col = _EXTRA_PAL[_extra_idx[0] % len(_EXTRA_PAL)]
                _extra_idx[0] += 1
                _COL[key] = col
                _LBL[key] = name
                e_custom.delete(0, "end")
                _rebuild_type_radios()
                sel_type.set(key)

            e_custom.bind("<Return>", _add_custom_type)
            self._btn(add_row, "+", _add_custom_type, accent=True).pack(side="left")

            tk.Frame(right_fr, bg=CLR["border"], height=1).pack(fill="x", padx=10, pady=(4, 0))
            tk.Label(right_fr, text="추가된 구역", font=FONT_SMALL,
                     fg=CLR["subtext"], bg=CLR["panel"]).pack(
                anchor="w", padx=10, pady=(4, 2))

            list_inner_fr = tk.Frame(right_fr, bg=CLR["panel"])
            list_inner_fr.pack(fill="x", padx=8)

            def _refresh_list():
                for w_ in list_inner_fr.winfo_children():
                    w_.destroy()
                for r in regions:
                    row2 = tk.Frame(list_inner_fr, bg=CLR["panel"])
                    row2.pack(anchor="w", fill="x", pady=1, padx=2)
                    tk.Frame(row2, bg=_COL.get(r["type"], "#555"),
                             width=8, height=8).pack(side="left", padx=(0, 4))
                    tag  = r.get("label") or _LBL.get(r["type"], r["type"])
                    disp = r.get("text", "")[:16] or "…"
                    lbl2 = tk.Label(row2, text=f"{tag}: {disp}",
                                    font=("Consolas", 8),
                                    fg=CLR["text"], bg=CLR["panel"], cursor="hand2")
                    lbl2.pack(side="left")
                    lbl2.bind("<Button-1>", lambda e, _r=r: _classify(_r))

                    def _del_r(_r=r):
                        for cid in _r.get("cids", []):
                            canvas.delete(cid)
                        if _r in regions:
                            regions.remove(_r)
                        _refresh_list()
                        _render()

                    tk.Button(row2, text="×", command=_del_r,
                              font=("Consolas", 9, "bold"),
                              fg=CLR["ng"], bg=CLR["panel"], bd=0,
                              cursor="hand2", relief="flat",
                              activebackground=CLR["panel"],
                              activeforeground="#ff6666").pack(side="right", padx=4)

            tk.Label(right_fr,
                     text="• 왼쪽 드래그: 구역 추가\n• 구역 클릭: 편집/삭제\n"
                          "• 스크롤: 확대/축소\n• 오른쪽 드래그: 화면 이동",
                     font=("Consolas", 8), fg=CLR["subtext"], bg=CLR["panel"],
                     justify="left").pack(anchor="w", padx=10, pady=6)

            # ── 저장 ─────────────────────────────
            def _save():
                name = e_name.get().strip() or "NEW_TEMPLATE"
                if not regions:
                    messagebox.showwarning("알림", "구역을 하나 이상 추가해주세요.", parent=win)
                    return
                new_data = {
                    "template_name": name,
                    "img_size":      [iw, ih],
                    "regions": [
                        {"label": r.get("label", _LBL.get(r["type"], r["type"])),
                         "type":  r["type"],
                         "rect":  list(r["rect"]),
                         "text":  r.get("text", "")}
                        for r in regions
                    ],
                }
                _init_file = (os.path.basename(self.tmpl.path)
                              if is_edit and self.tmpl.path else f"{name}.json")
                _init_dir  = (os.path.dirname(self.tmpl.path)
                              if is_edit and self.tmpl.path else self._app_dir)
                path = filedialog.asksaveasfilename(
                    defaultextension=".json",
                    filetypes=[("JSON 템플릿", "*.json")],
                    initialfile=_init_file,
                    initialdir=_init_dir or self._app_dir,
                    title="템플릿 저장", parent=win)
                if not path:
                    return
                # 참조 이미지 저장 (추적용)
                ref_name = os.path.splitext(os.path.basename(path))[0] + "_ref.jpg"
                ref_path = os.path.join(os.path.dirname(path), ref_name)
                cv2.imwrite(ref_path, bgr_orig)
                new_data["ref_image"] = ref_name
                self.tmpl.data = new_data
                self.tmpl.save(path)
                self.tmpl._load_ref_image()   # ORB 특징점 즉시 계산
                self._lbl_tmpl.configure(text=f"[ {name} ]", fg=CLR["ok"])
                action = "수정" if is_edit else "생성"
                self._status(f"템플릿 {action}: {os.path.basename(path)}")
                win.destroy()
                self._refresh_tmpl_overlay()

            bot_fr = tk.Frame(right_fr, bg=CLR["panel"])
            bot_fr.pack(side="bottom", fill="x", padx=10, pady=10)
            self._btn(bot_fr, "저장", _save, accent=True).pack(fill="x", pady=(0, 4))
            self._btn(bot_fr, "닫기", win.destroy).pack(fill="x")

            def _reset_zoom():
                zoom[0] = 1.0;  pan[0] = 0;  pan[1] = 0
                _render()

            self._btn(toolbar, "초기화", _reset_zoom).pack(side="right", padx=(4, 0))
            tk.Label(toolbar,
                     text="스크롤: 확대/축소   오른쪽 드래그: 이동   왼쪽 드래그: 구역 지정   구역 클릭: 편집",
                     font=FONT_SMALL, fg=CLR["subtext"], bg=CLR["bg"]).pack(
                side="left", padx=8)

        try:
            # 편집 모드: 영역이 그려진 원본 기준 이미지(ref_bgr)를 배경으로 사용해야
            # 영역이 라벨과 정확히 정렬된다. 현재 카메라 프레임은 라벨 위치/크기가
            # 달라 영역이 어긋나 보이므로 사용하지 않는다.
            edit_bg = None
            if (initial_data is not None
                    and getattr(self.tmpl, "ref_bgr", None) is not None):
                edit_bg = self.tmpl.ref_bgr.copy()
            if edit_bg is not None:
                _open_editor(edit_bg)
            elif self.current_bgr is not None:
                _open_editor(self.current_bgr.copy())
            else:
                path = filedialog.askopenfilename(
                    title="라벨 이미지 선택",
                    filetypes=[("이미지", "*.png *.jpg *.jpeg *.bmp *.tiff"),
                               ("모든 파일", "*.*")],
                )
                if not path:
                    return
                bgr = cv2.imdecode(np.fromfile(path, dtype=np.uint8), cv2.IMREAD_COLOR)
                if bgr is None:
                    messagebox.showerror("오류", "이미지를 열 수 없습니다.")
                    return
                _open_editor(bgr)
        except Exception:
            import traceback
            messagebox.showerror("오류", traceback.format_exc())

    def _load_template(self):
        path = filedialog.askopenfilename(
            title="템플릿 파일 선택",
            initialdir=self._app_dir,
            filetypes=[("JSON 템플릿", "*.json"), ("모든 파일", "*.*")]
        )
        if not path:
            return
        try:
            self._stop_tracking()
            self.tmpl.load(path)
            self._lbl_tmpl.configure(text=f"[ {self.tmpl.name} ]", fg=CLR["ok"])
            n = len(self.tmpl.data.get("regions", []))
            self._status(f"템플릿 로드: {os.path.basename(path)}  ({n}개 구역)")
            if "regions" not in self.tmpl.data:
                messagebox.showwarning(
                    "구 버전 템플릿",
                    f"'{os.path.basename(path)}'은 구 버전 형식입니다.\n\n"
                    "v2.0부터 구역(regions) 기반 포맷으로 변경되어\n"
                    "이 파일에는 오버레이 구역 정보가 없습니다.\n\n"
                    "▶ '+ 새 템플릿' 버튼으로 구역을 새로 지정해주세요."
                )
                return
            if n == 0:
                messagebox.showwarning(
                    "빈 템플릿",
                    "템플릿에 구역이 없습니다.\n'✏ 편집' 버튼으로 구역을 추가해주세요."
                )
                return
            self._refresh_tmpl_overlay()
        except Exception as e:
            messagebox.showerror("오류", f"템플릿 로드 실패:\n{e}")

    def _refresh_tmpl_overlay(self):
        """현재 이미지에 템플릿 구역 윤곽선 표시 + 추적 스레드 재시작."""
        if not self.tmpl.loaded:
            return
        regs = self.tmpl.data.get("regions", [])
        if not regs:
            self._status("⚠ 템플릿에 구역 없음 — '✏ 편집'으로 구역을 추가하세요")
            return
        self._stop_tracking()
        if self.cam_running:
            self._start_tracking()
        if self.current_bgr is not None:
            self._show_frame(self.current_bgr, tmpl_preview=True)
        else:
            track_hint = "  (추적 기능 포함)" if self.tmpl.has_tracking else ""
            self._status(f"템플릿 로드됨 ({len(regs)}개 구역){track_hint} — 카메라 또는 이미지를 열면 구역이 표시됩니다")

    def _edit_template(self):
        """로드된 템플릿을 캔버스 편집기로 열기."""
        if not self.tmpl.loaded:
            messagebox.showinfo("알림", "먼저 템플릿을 불러오거나 새로 만들어주세요.")
            return
        self._template_wizard(initial_data=self.tmpl.data)

    def _clear_template(self):
        """불러온 템플릿 해제."""
        if not self.tmpl.loaded:
            return
        self._stop_tracking()
        self.tmpl.data    = None
        self.tmpl.path    = ""
        self.tmpl.ref_bgr = None
        self.tmpl._orb_kp  = None
        self.tmpl._orb_des = None
        self._lbl_tmpl.configure(text="[ 템플릿 없음 ]", fg=CLR["warn"])
        if self.current_bgr is not None:
            self._show_frame(self.current_bgr)
        self._status("템플릿 해제됨")

    # ── 이력 ─────────────────────────────────────
    def _show_history(self):
        if not self.history:
            messagebox.showinfo("이력", "검사 이력이 없습니다.")
            return
        win = tk.Toplevel(self)
        win.title("검사 이력")
        win.configure(bg=CLR["bg"])
        win.geometry("700x440")

        total  = len(self.history)
        ok_cnt = sum(1 for h in self.history if h["verdict"] == "OK")
        stat_fr = tk.Frame(win, bg=CLR["panel"])
        stat_fr.pack(fill="x", padx=8, pady=(8, 0))
        tk.Label(stat_fr, text=f"총 {total}건", font=FONT_SMALL,
                 fg=CLR["text"],  bg=CLR["panel"]).pack(side="left", padx=10, pady=4)
        tk.Label(stat_fr, text=f"OK {ok_cnt}", font=FONT_SMALL,
                 fg=CLR["ok"],   bg=CLR["panel"]).pack(side="left", padx=6)
        tk.Label(stat_fr, text=f"NG {total - ok_cnt}", font=FONT_SMALL,
                 fg=CLR["ng"],   bg=CLR["panel"]).pack(side="left", padx=6)

        tree_fr = tk.Frame(win, bg=CLR["bg"])
        tree_fr.pack(fill="both", expand=True, padx=8, pady=4)
        cols = ("시각", "판정", "템플릿", "구역", "결과", "인식값")
        tv   = ttk.Treeview(tree_fr, columns=cols, show="headings")
        sb   = ttk.Scrollbar(tree_fr, command=tv.yview)
        tv.configure(yscrollcommand=sb.set)
        widths = {"시각": 140, "판정": 55, "템플릿": 110, "구역": 90,
                  "결과": 50, "인식값": 200}
        for c in cols:
            tv.heading(c, text=c)
            tv.column(c, width=widths.get(c, 80), anchor="center")

        for entry in reversed(self.history):
            for r in entry.get("regions", []):
                tv.insert("", "end", values=(
                    entry["timestamp"], entry["verdict"], entry["template"],
                    r["label"], "OK" if r["ok"] else "NG", r["value"][:40],
                ), tags=(entry["verdict"],))

        tv.tag_configure("OK", foreground=CLR["ok"])
        tv.tag_configure("NG", foreground=CLR["ng"])
        sb.pack(side="right", fill="y")
        tv.pack(side="left",  fill="both", expand=True)
        self._btn(win, "닫기", win.destroy).pack(pady=(0, 8))

    # ── 카메라 선택 다이얼로그 ────────────────────
    def _show_cam_dialog(self):
        win = tk.Toplevel(self)
        win.title("카메라 연결")
        win.configure(bg=CLR["bg"])
        win.geometry("480x220")
        win.resizable(False, False)
        win.grab_set()

        usb_fr = tk.LabelFrame(win, text=" 🖥  USB 카메라 ",
                               font=FONT_SMALL, fg=CLR["accent"],
                               bg=CLR["bg"], bd=1)
        usb_fr.pack(fill="x", padx=14, pady=(12, 6))

        usb_row = tk.Frame(usb_fr, bg=CLR["bg"])
        usb_row.pack(fill="x", padx=8, pady=8)

        _cam_map = {}
        lb = tk.Listbox(usb_row, font=FONT_SMALL, bg=CLR["btn"], fg=CLR["text"],
                        selectbackground=CLR["accent"], selectforeground="#FFFFFF",
                        relief="flat", height=4, activestyle="none", bd=0, width=40)
        lb.pack(side="left", fill="x", expand=True, padx=(0, 6))
        lb.insert("end", "  ( 스캔 버튼을 눌러 검색 )")

        btn_col = tk.Frame(usb_row, bg=CLR["bg"])
        btn_col.pack(side="left")

        def on_usb_connect():
            sel = lb.curselection()
            if not sel or sel[0] not in _cam_map:
                messagebox.showwarning("경고", "카메라를 선택해주세요.", parent=win)
                return
            cam_idx, backend = _cam_map[sel[0]]
            rs = self._res_var.get().replace("×", "x")
            w, h = map(int, rs.split("x"))
            self._cam_fmt = self._fmt_var.get()
            self._cam_res = (w, h)
            win.destroy()
            self._stop_camera()
            self._stop_video()
            self._start_camera(cam_idx, self._cam_fmt, (w, h), backend)

        lb.bind("<Double-Button-1>", lambda e: on_usb_connect())

        def scan_usb():
            btn_scan.configure(state="disabled", text="스캔 중...")
            lb.delete(0, "end")
            _cam_map.clear()
            lb.insert("end", "  스캔 중...")

            def _bg():
                ps_names = App._get_ps_cam_names()
                found = [];  ps_i = 0
                for idx in range(6):
                    result = [None]   # (backend, label_suffix)
                    def try_open(i=idx, r=result):
                        for bk, bk_name in ((cv2.CAP_DSHOW, "DSHOW"), (cv2.CAP_MSMF, "MSMF")):
                            try:
                                c = cv2.VideoCapture(i, bk)
                                if not c.isOpened():
                                    c.release(); continue
                                if bk == cv2.CAP_MSMF:
                                    c.set(cv2.CAP_PROP_FOURCC,
                                          cv2.VideoWriter_fourcc(*'MJPG'))
                                import time as _t; _t.sleep(0.3)
                                ret, _ = c.read()
                                c.release()
                                if ret:
                                    r[0] = (bk, "" if bk_name == "DSHOW" else f" [{bk_name}]")
                                    return
                            except Exception:
                                pass
                    t = threading.Thread(target=try_open, daemon=True)
                    t.start();  t.join(timeout=4.0)
                    if result[0]:
                        bk, suffix = result[0]
                        name = ps_names.get(ps_i, f"CAM {idx}")
                        found.append((idx, bk, f"  {idx} — {name}{suffix}"))
                        ps_i += 1

                def _update():
                    lb.delete(0, "end");  _cam_map.clear()
                    if found:
                        for li, (ci, bk, lbl) in enumerate(found):
                            lb.insert("end", lbl);  _cam_map[li] = (ci, bk)
                        # 내장 카메라가 아닌 첫 번째 항목 자동 선택
                        auto = next(
                            (li for li, (_, _, lbl) in enumerate(found)
                             if not any(k in lbl.lower() for k in ("lg", "built", "internal", "integrated"))),
                            0)
                        lb.selection_set(auto)
                        lb.see(auto)
                    else:
                        lb.insert("end", "  카메라 없음")
                    btn_scan.configure(state="normal", text="🔍 스캔")
                win.after(0, _update)

            threading.Thread(target=_bg, daemon=True).start()

        btn_scan = self._btn(btn_col, "🔍 스캔", scan_usb)
        btn_scan.pack(pady=(0, 4))
        self._btn(btn_col, "🔗 연결", on_usb_connect, accent=True).pack()

        bot = tk.Frame(win, bg=CLR["bg"])
        bot.pack(fill="x", padx=14, pady=(2, 10))
        self._btn(bot, "취소", win.destroy).pack(side="left")

    @staticmethod
    def _get_ps_cam_names():
        try:
            import subprocess
            script = (
                'Get-PnpDevice -PresentOnly | '
                'Where-Object { $_.Class -in @("Camera","Image") -and $_.Status -eq "OK" } | '
                'Sort-Object InstanceId | Select-Object -ExpandProperty FriendlyName'
            )
            r = subprocess.run(
                ['powershell', '-NoProfile', '-NonInteractive', '-Command', script],
                capture_output=True, text=True, timeout=5
            )
            names = [ln.strip() for ln in r.stdout.splitlines() if ln.strip()]
            return {i: n for i, n in enumerate(names)}
        except Exception:
            return {}

    # ── 자동 검사 ─────────────────────────────────
    def _toggle_auto_inspect(self):
        self._auto_inspect = not self._auto_inspect
        if self._auto_inspect:
            try:
                self._auto_interval = max(2, int(self._auto_spin.get()))
            except ValueError:
                self._auto_interval = 5
            self._btn_auto.configure(bg=CLR["btn_on"])
            self._status(f"자동 검사 시작 — {self._auto_interval}초 간격")
            self._auto_inspect_loop()
        else:
            self._btn_auto.configure(bg=CLR["btn"])
            if self._auto_after_id:
                self.after_cancel(self._auto_after_id)
                self._auto_after_id = None
            self._status("자동 검사 중지됨")

    def _auto_inspect_loop(self):
        if not self._auto_inspect:
            return
        if self.cam_running and self.current_bgr is not None:
            if str(self._btn_snap.cget("state")) == "normal":
                self._inspect()
        self._auto_after_id = self.after(
            self._auto_interval * 1000, self._auto_inspect_loop)

    # ── 설정 저장 / 복원 ─────────────────────────
    def _save_config(self):
        idx = self._current_cam_idx
        cfg = {
            "fmt":         self._fmt_var.get(),
            "res":         self._res_var.get(),
            "rotate":      self._rotate,
            "cam_type":    "index",
            "cam_idx":     idx if isinstance(idx, int) else 0,
            "cam_backend": int(getattr(self, "_current_cam_backend", cv2.CAP_DSHOW)),
            "tmpl_path":   self.tmpl.path if self.tmpl.loaded else "",
        }
        try:
            with open(self._config_path, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _load_config(self):
        self._last_cam_cfg = {}
        try:
            with open(self._config_path, encoding="utf-8") as f:
                cfg = json.load(f)
            self._last_cam_cfg = cfg   # 위젯 갱신 실패와 무관하게 카메라 설정 보존
            self._fmt_var.set(cfg.get("fmt", "YUY2"))
            self._res_var.set(cfg.get("res", "640×480"))
            self._rotate = cfg.get("rotate", 0)
            self._lbl_rotate.configure(text=f"{self._rotate}°")
            self._last_cam_cfg = cfg
            # 마지막 템플릿 자동 로드
            tmpl_path = cfg.get("tmpl_path", "")
            if tmpl_path and os.path.exists(tmpl_path):
                try:
                    self.tmpl.load(tmpl_path)
                    if "regions" not in self.tmpl.data:
                        # 구 버전 형식 — 사용 불가, 로드 취소
                        self.tmpl.data = None
                        self.tmpl.path = ""
                        self.after(400, lambda: self._status(
                            f"⚠ 구 버전 템플릿 형식 ({os.path.basename(tmpl_path)}) — "
                            "'+ 새 템플릿'으로 새로 만들어주세요"
                        ))
                    else:
                        self._lbl_tmpl.configure(text=f"[ {self.tmpl.name} ]", fg=CLR["ok"])
                except Exception:
                    pass
        except Exception:
            pass
        # 이전 카메라 자동 연결
        self._auto_connect_camera()

    def _auto_connect_camera(self):
        """설정에 저장된 카메라로 자동 연결."""
        cfg = self._last_cam_cfg
        if not cfg:
            return
        cam_type    = cfg.get("cam_type", "")
        cam_url     = cfg.get("cam_url", "")
        cam_idx     = cfg.get("cam_idx", 0)
        cam_backend = cfg.get("cam_backend", cv2.CAP_DSHOW)
        fmt         = cfg.get("fmt", "YUY2")
        res_str     = cfg.get("res", "640×480")
        try:
            w, h = map(int, res_str.replace("×", "x").split("x"))
        except ValueError:
            w, h = 640, 480
        self._cam_fmt = fmt
        self._cam_res = (w, h)
        if cam_type == "index":
            self._status(f"이전 카메라 자동 연결 중... CAM {cam_idx}")
            self.after(1200, lambda ci=cam_idx, f=fmt, r=(w, h), b=cam_backend:
                       self._start_camera(ci, f, r, b))

    # ── 상태 바 / 종료 ────────────────────────────
    def _status(self, msg):
        self._statusbar.configure(text=f"  {msg}")

    def _on_close(self):
        if self._auto_after_id:
            self.after_cancel(self._auto_after_id)
        self._save_config()
        self._stop_camera()
        self._stop_video()
        self.destroy()


# ──────────────────────────────────────────────
if __name__ == "__main__":
    import ctypes as _ctypes

    app = App()
    app.geometry("1100x680")

    def _move_to_front():
        app.deiconify()
        app.update()
        _title = "Optical Transceiver Label Inspector  v2.0"
        hwnd = _ctypes.windll.user32.FindWindowW(None, _title)
        if hwnd:
            _W, _H = 1100, 680
            # SM_CXSCREEN=0, SM_CYSCREEN=1 → 주 모니터 해상도
            _sw = _ctypes.windll.user32.GetSystemMetrics(0)
            _sh = _ctypes.windll.user32.GetSystemMetrics(1)
            _x  = max(0, (_sw - _W) // 2)
            _y  = max(0, (_sh - _H) // 2)
            _ctypes.windll.user32.SetWindowPos(hwnd, 0, _x, _y, _W, _H, 0x0044)
            _ctypes.windll.user32.SetForegroundWindow(hwnd)

    app.after(800, _move_to_front)
    app.mainloop()
